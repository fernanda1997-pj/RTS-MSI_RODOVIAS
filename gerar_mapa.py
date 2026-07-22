import folium
from folium.plugins import Fullscreen, MousePosition, MarkerCluster
from branca.element import MacroElement, Template
import geopandas as gpd
import pandas as pd
from pathlib import Path
import warnings
import base64
import re
import sys
import math
import unicodedata
import html as _html

warnings.filterwarnings('ignore')  # Ignorar avisos do geopandas

# --------------------------------------------------- pontos críticos
# Planilha de controle: uma aba por região, com o status de cada ponto mês a
# mês e o link da foto de cada mês (hospedadas no Drive, permanecem privadas).
PLANILHA_PC = Path(__file__).parent / 'pontos criticos' / 'Controle Pontos Críticos .xlsx'

# Ordem cronológica do acompanhamento (o ciclo vai de setembro a julho)
MESES = ['Setembro', 'Outubro', 'Novembro', 'Dezembro', 'Janeiro',
         'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho']

# Status do ponto -> (rótulo exibido, cor)
STATUS_PC = {
    'critico':      ('Crítico',      '#DC2626'),
    'em execucao':  ('Em execução',  '#F59E0B'),
    'recuperado':   ('Recuperado',   '#059669'),
}
STATUS_PC_OUTRO = ('Sem registro', '#94A3B8')

# Ordem de exibição dos status na árvore (do mais grave ao resolvido)
ORDEM_STATUS_PC = ['Crítico', 'Em execução', 'Recuperado', 'Sem registro']

# Classe CSS por status, para o halo/brilho do marcador no mapa
CLASSE_STATUS_PC = {
    '#DC2626': 'pc-c1',   # crítico (vermelho)
    '#F59E0B': 'pc-c2',   # em execução (âmbar)
    '#059669': 'pc-c3',   # recuperado (verde)
    '#94A3B8': 'pc-c0',   # sem registro (cinza)
}

# ---------------------------------------------------------------- identidade
COR_MARCA = '#1E3A72'  # azul-marinho da RTA (extraído do logo)

# Cores por tipo de camada (dados por região)
COR_REGIAO = '#1E3A72'   # fallback e cor do símbolo genérico de região

# Uma cor por região. Todas dessaturadas e na mesma faixa de luminosidade
# (por isso formam família e não competem com as rodovias), mas com matizes
# bem espaçados (por isso dá para diferenciar uma da outra). São polígonos
# de contexto desenhados a 35% de opacidade: o contorno é que carrega a cor.
CORES_REGIAO = {
    'R1':  '#7391B5',   # azul-acinzentado
    'R2':  '#74A88E',   # verde-água
    'R3':  '#A88BB5',   # lilás
    'R11': '#C49A6C',   # caramelo
    'R12': '#B57A8A',   # rosa-vinho
    'R13': '#8A9E6B',   # verde-oliva
}
# Se aparecer uma região nova, recebe uma cor daqui de forma estável
PALETA_REGIAO_EXTRA = ['#6E9BA8', '#B58F7A', '#9A8FB5', '#8FA87E', '#AE8E9E']


def cor_regiao(rid):
    """Cor do polígono de uma região (estável entre execuções)."""
    if rid in CORES_REGIAO:
        return CORES_REGIAO[rid]
    mm = re.match(r'R(\d+)$', rid)
    i = int(mm.group(1)) if mm else sum(ord(c) for c in rid)
    return PALETA_REGIAO_EXTRA[i % len(PALETA_REGIAO_EXTRA)]
# Pontos SRE: nós de referência nos extremos dos trechos. Ficam neutros
# (miolo branco + anel marinho) para marcar sem competir com as rodovias,
# que já usam a paleta colorida da situação.
COR_SRE = '#1E3A72'
COR_CRITICO = '#ef2b2b'

# Cores das camadas de contexto (estaduais)
COR_ESTADO = '#94A3B8'  # slate: legível no satélite E no painel branco
COR_HIDRO = '#38BDF8'   # azul-céu

# Situação dos trechos (códigos oficiais do SRE) -> (descrição, cor).
# As cores são agrupadas por FAMÍLIA para o mapa se ler à distância:
#   verdes  = duplicadas | azuis = pavimentadas simples
#   laranjas = em obras  | marrons = não pavimentadas | cinza = planejada
# A ordem do dicionário vai do melhor padrão ao mais precário (e define a
# ordem de exibição na legenda).
SITUACAO_INFO = {
    'DUP': ('Duplicada',                     '#047857'),  # verde escuro
    'PDU': ('Duplicada em perímetro urbano', '#10B981'),  # verde claro
    'PPS': ('Pavimentada pista simples',     '#2563EB'),  # azul
    'PSU': ('Simples urbana',                '#60A5FA'),  # azul claro
    'EOD': ('Em obras de duplicação',        '#FB923C'),  # laranja
    'EOP': ('Em obras de pavimentação',      '#F59E0B'),  # âmbar
    'IMP': ('Implantada',                    '#A16207'),  # ocre
    'LEN': ('Leito natural',                 '#B45309'),  # marrom (terra)
    'PLA': ('Planejada',                     '#94A3B8'),  # cinza
}
DESC_OUTRA = 'Não classificada'
COR_SIT_OUTRA = '#CBD5E1'  # fallback para código fora da tabela do SRE

# Hidrografia: nuordemcda 1 = curso principal (Rio Tocantins, bacia de
# ~940 mil km²); ordens altas são riachos. Filtramos aos principais para o
# mapa não estourar de tamanho (base completa tem 77.795 feições).
HIDRO_ORDEM_MAX = 3
HIDRO_SIMPLIFY = 0.0005  # ~55 m

TIPO_META = {
    'regiao':   ('Formato da Região', COR_REGIAO,  'poli'),
    'trechos':  ('Trechos',           COR_MARCA,   'multi'),
    'sre':      ('Pontos SRE',        COR_SRE,     'no'),
    'criticos': ('Pontos críticos',   COR_CRITICO, 'crit'),
}
TIPO_ORDEM = ['regiao', 'trechos', 'sre', 'criticos']

# Inventário de elementos (drenagem, seguranca e OAEs). Um shapefile por tipo
# e por regiao, nomeado {TIPO}_R{n} (ex.: OAE_R11, MEIO_FIO_R11).
# (chave no nome do arquivo, rótulo, cor, forma do símbolo)
# Cores em tríade (matizes ~120° de distância): o máximo de contraste
# possível entre três, e nenhuma repete as cores já usadas no mapa.
INVENTARIO_META = [
    ('OAE',      'Pontes (OAE)',      '#FACC15', 'ponto'),   # obra de arte especial
    ('OAC',      'Bueiros (OAC)',     '#9333EA', 'ponto'),   # obra de arte corrente
    ('DESCIDA',  "Descida d'água",    '#0EA5E9', 'ponto'),
]

# Camadas de linha do inventário, desativadas a pedido. Os shapefiles seguem
# na pasta e são ignorados em silêncio. Para reativar, mova a entrada de volta
# para INVENTARIO_META (o código de linhas continua pronto):
#   ('VALETA',   'Valetas',   '#15803D', 'linha')
#   ('SARJETA',  'Sarjetas',  '#CA8A04', 'linha')
#   ('MEIO_FIO', 'Meio-fio',  '#64748B', 'linha')
#   ('DEFENSA',  'Defensas',  '#9333EA', 'linha')
INVENTARIO_DESATIVADOS = {'VALETA', 'SARJETA', 'MEIO_FIO', 'DEFENSA'}

# Erros de digitação já vistos nos nomes de arquivo. Aceitamos o apelido (para
# não perder a camada em silêncio) mas avisamos no console para corrigir.
INVENTARIO_APELIDOS = {'AOE': 'OAE'}

INVENTARIO_TIPOS = {m[0] for m in INVENTARIO_META}


def geom_valida(geom):
    """
    True se a geometria existe e tem coordenadas finitas dentro do globo.
    Filtra pontos com lat/lon NaN ou fora de faixa (que travam o Leaflet ao
    calcular limites — foi o caso de um ponto da Descida d'água da R11).
    """
    if geom is None or geom.is_empty:
        return False
    try:
        minx, miny, maxx, maxy = geom.bounds
    except Exception:
        return False
    if not all(math.isfinite(v) for v in (minx, miny, maxx, maxy)):
        return False
    return -180 <= minx <= 180 and -180 <= maxx <= 180 and -90 <= miny <= 90 and -90 <= maxy <= 90


def colunas_uteis(gdf):
    """
    Colunas que valem a pena mostrar no popup do inventário. Descarta lixo de
    join do ArcGIS (Field13, sufixos _1, OBJECTID) e coordenadas soltas
    (X/Y, que a posição no mapa já dá).
    """
    fora = re.compile(r'^(field\d+|objectid|rid|meas|distance|[xy]_?(inicio|final)?|'
                      r'.*_\d+)$', re.IGNORECASE)
    uteis = []
    for c in gdf.columns:
        if c == 'geometry':
            continue
        if fora.match(str(c)):
            continue
        uteis.append(c)
    return uteis

# Arquivos redundantes/brutos que não devem entrar no mapa
IGNORAR = {
    'esatado_tocantins',     # duplicata (erro de digitação) de estado_tocantins
    'hidrografia_estado_1',  # duplicata de hidrografia_estado
    'hidrografia',           # base bruta: extrapola o TO (usar a recortada)
}


def limpar_atributos(gdf):
    """
    Remove quebras de linha (\\r\\n) e espaços extras dos campos de texto.
    Sem isso, 'PSU' e 'PSU\\r\\n' viram categorias diferentes na legenda.

    Atenção ao teste de tipo: o geopandas 1.1 lê texto como dtype 'str', não
    'object'. A versão anterior checava 'dtype != object' e por isso pulava
    TODAS as colunas de texto — a função não limpava nada. Aqui invertemos:
    limpa tudo que não for número, data ou geometria.
    """
    for col in gdf.columns:
        if col == 'geometry' or gdf[col].dtype.name == 'geometry':
            continue
        if (pd.api.types.is_numeric_dtype(gdf[col])
                or pd.api.types.is_datetime64_any_dtype(gdf[col])
                or pd.api.types.is_bool_dtype(gdf[col])):
            continue
        gdf[col] = (gdf[col].astype(str)
                    .str.replace(r'[\r\n\t]+', ' ', regex=True)
                    .str.strip()
                    .replace({'nan': None, 'None': None, '': None}))
    return gdf


def validar_nomenclatura_sre(gdf, nome_camada):
    """Controle de Qualidade: alerta sobre SRE vazio ou duplicado."""
    colunas_sre = [c for c in gdf.columns if 'SRE' in str(c).upper()]
    if not colunas_sre:
        return
    for col in colunas_sre:
        serie = gdf[col]
        total = len(serie)
        vazios = serie.isna() | serie.astype(str).str.strip().isin(['', 'None', 'nan'])
        n_vazios = int(vazios.sum())
        if n_vazios > 0:
            print(f"  [QUALIDADE] '{nome_camada}' -> '{col}': "
                  f"{n_vazios} de {total} registro(s) com SRE VAZIO.")
        preenchidos = serie[~vazios]
        duplicados = preenchidos[preenchidos.duplicated(keep=False)]
        if len(duplicados) > 0:
            print(f"  [QUALIDADE] '{nome_camada}' -> '{col}': "
                  f"SRE DUPLICADO(S): {', '.join(sorted(set(duplicados.astype(str))))}")
        if n_vazios == 0 and len(duplicados) == 0:
            print(f"  [QUALIDADE] '{nome_camada}' -> '{col}': OK ({total} registros).")


def classificar(nome):
    """
    Identifica o tipo da camada pelo nome do arquivo.

    Tolerante a acento e a separador: 'Pontos críticos', 'Pontos_Criticos' e
    'PONTOS CRITICOS' caem todos em 'criticos'.
    """
    if nome.lower() in IGNORAR:
        return 'skip'
    # Inventário: {TIPO}_R{n} (ex.: OAE_R11, DESCIDA_R11)
    mm = re.match(r'^(.+)_R\d+$', nome, re.IGNORECASE)
    if mm:
        t = mm.group(1).upper()
        if t in INVENTARIO_TIPOS or t in INVENTARIO_APELIDOS:
            return 'inventario'
        if t in INVENTARIO_DESATIVADOS:
            return 'skip'
    n = _sem_acento(nome).replace('_', ' ').replace('-', ' ')
    if n.startswith('hidrografia'):        # antes de 'estado' (hidrografia_estado)
        return 'hidrografia'
    if 'estado' in n and 'tocantins' in n:
        return 'estado'
    if 'pontos criticos' in n or 'ponto critico' in n:
        return 'criticos'
    if 'pontos sre' in n:
        return 'sre'
    if 'trechos' in n:
        return 'trechos'
    if 'regi' in n:
        return 'regiao'
    return None


def coluna_situacao(gdf):
    for c in gdf.columns:
        if 'SITUA' in str(c).upper():
            return c
    return None


def ordenar_situacoes(codigos):
    """Ordem fixa (a do SITUACAO_INFO) para comparar regiões lado a lado."""
    ordem = list(SITUACAO_INFO)
    return sorted(codigos, key=lambda c: (ordem.index(c) if c in ordem else 99, c))


def info_situacao(codigo):
    """Devolve (descrição, cor) de um código de situação do SRE."""
    return SITUACAO_INFO.get(codigo, (DESC_OUTRA, COR_SIT_OUTRA))


def _sem_acento(s):
    return unicodedata.normalize('NFD', str(s or '')).encode('ascii', 'ignore').decode().lower().strip()


def _detectar_mes(cabecalho):
    """
    Descobre a que mês uma coluna se refere e se é de foto.

    Os cabeçalhos da planilha são inconsistentes entre as abas:
    'Mapas de Out', 'Mapa de Dezembro', 'MapaAbril', 'mapa Maio', 'Mapa fev'...
    Atenção ao \\s+ em '^de\\s+': sem ele o 'de' de 'DEzembro' seria removido
    e o mês se perderia silenciosamente.
    """
    n = _sem_acento(cabecalho)
    eh_foto = n.startswith('mapa')
    n2 = re.sub(r'^mapas?\s*', '', n)
    n2 = re.sub(r'^de\s+', '', n2).strip()
    for m in MESES:
        if n2.startswith(_sem_acento(m)[:3]):
            return m, eh_foto
    return None, eh_foto


def _status_pc(valor):
    """Normaliza o status ('Em execução' e 'Em Execução' são o mesmo)."""
    return STATUS_PC.get(_sem_acento(valor), STATUS_PC_OUTRO)


def carregar_controle_pontos():
    """
    Lê a planilha de controle e devolve {(regiao, ponto): dados}.
    Sem a planilha (ou sem openpyxl), devolve {} e o mapa segue sem as fichas.
    """
    if not PLANILHA_PC.exists():
        print(f"Aviso: planilha não encontrada em {PLANILHA_PC.name}. "
              f"Os pontos críticos entram sem ficha.")
        return {}
    try:
        import openpyxl
    except ImportError:
        print("Aviso: openpyxl não instalado (pip install openpyxl). "
              "Os pontos críticos entram sem ficha.")
        return {}

    wb = openpyxl.load_workbook(PLANILHA_PC)
    dados = {}
    for ws in wb.worksheets:
        achado = re.search(r'(\d+)', ws.title)
        if not achado:
            continue
        rid = 'R' + achado.group(1)
        hdr = [c.value for c in ws[1]]

        # 'Status Final / Situação' na maioria das abas; 'Situação Atual' na R13
        i_final = next((i for i, h in enumerate(hdr)
                        if h and ('status final' in _sem_acento(h)
                                  or 'situacao atual' in _sem_acento(h))), None)
        col_status, col_foto = {}, {}
        for i, h in enumerate(hdr):
            mes, eh_foto = _detectar_mes(h)
            if mes:
                (col_foto if eh_foto else col_status)[mes] = i

        for row in ws.iter_rows(min_row=2):
            if not row[0].value:
                continue
            try:
                ponto = int(float(row[0].value))
            except (TypeError, ValueError):
                continue
            historico = []
            for mes in MESES:
                i_s, i_f = col_status.get(mes), col_foto.get(mes)
                st = row[i_s].value if i_s is not None else None
                st = str(st).strip() if st and str(st).strip() not in ('·', '', '-') else None
                foto = row[i_f].hyperlink.target if (i_f is not None and row[i_f].hyperlink) else None
                if st or foto:
                    historico.append({'mes': mes, 'status': st, 'foto': foto})
            dados[(rid, ponto)] = {
                'rodovia': row[1].value,
                'trecho': row[2].value,
                'final': row[i_final].value if i_final is not None else None,
                'historico': historico,
            }
    print(f"Planilha de controle: {len(dados)} pontos, "
          f"{sum(1 for d in dados.values() for h in d['historico'] if h['foto'])} fotos.")
    return dados


def status_atual(info):
    """Status mais recente do ponto (último mês com registro)."""
    for h in reversed(info.get('historico', [])):
        if h['status']:
            return h['status']
    return None


def ficha_ponto_critico(rid, ponto, info):
    """Monta o HTML do popup de um ponto crítico."""
    e = _html.escape
    rotulo, cor = _status_pc(status_atual(info))
    partes = [
        '<div class="pc">',
        '<div class="pc-top">',
        f'<span class="pc-num">Ponto {ponto:03d}</span>',
        f'<span class="pc-tag" style="background:{cor}">{e(rotulo.upper())}</span>',
        '</div>',
    ]
    sub = ' · '.join(x for x in [str(info.get('rodovia') or '').strip(),
                                 str(info.get('trecho') or '').strip()] if x)
    if sub:
        partes.append(f'<div class="pc-sub">{e(sub)}</div>')

    if info.get('historico'):
        partes.append('<div class="pc-hist">')
        for h in info['historico']:
            r_h, c_h = _status_pc(h['status']) if h['status'] else ('—', '#CBD5E1')
            # O link abre a ficha do mês em PDF (mapa de localização + fotos),
            # hospedada no Drive — por isso "Ver mapa", e não "ver foto".
            foto = (f'<a class="pc-foto" href="{e(h["foto"])}" target="_blank" '
                    f'rel="noopener">Ver mapa</a>') if h['foto'] else '<span class="pc-nofoto">—</span>'
            partes.append(
                f'<div class="pc-linha"><span class="pc-mes">{e(h["mes"][:3])}</span>'
                f'<span class="pc-st" style="color:{c_h}">{e(r_h)}</span>{foto}</div>')
        partes.append('</div>')

    if info.get('final'):
        partes.append(f'<div class="pc-final">{e(str(info["final"]).strip())}</div>')
    partes.append('</div>')
    return ''.join(partes)


# EPSG métrico para medir comprimento (SIRGAS 2000 / UTM 22S, cobre o TO)
EPSG_METRICO = 31982


def normalizar_extensao(gdf, nome_arquivo):
    """
    Unifica o campo de extensão e devolve SEMPRE em quilômetros, na coluna
    EXT_KM.

    Duas inconsistências reais do inventário são tratadas aqui:

    1. NOME: cada região batizou o campo de um jeito -- 'EXTENSÃO',
       'EXTENSÃ_1', 'EXTENÇÃO', 'EXT_REAL', 'extenção'.

    2. UNIDADE: a R1 está em METROS e as outras cinco em QUILÔMETROS. Somar
       tudo sem converter daria ~1.037.000 km de malha em vez de ~6.652 km.

    A unidade é deduzida da GEOMETRIA (compara o atributo com o comprimento
    real medido), e não do nome do arquivo -- assim uma região nova em metros
    é tratada certo sozinha.
    """
    col = next((c for c in gdf.columns if 'EXT' in str(c).upper()), None)
    if col is None:
        return gdf
    try:
        valores = pd.to_numeric(gdf[col], errors='coerce')
        real_km = gdf.to_crs(EPSG_METRICO).geometry.length / 1000
        razao = (valores / real_km.replace(0, pd.NA)).median()
        if pd.isna(razao):
            return gdf
        if 900 < razao < 1100:          # atributo em metros
            gdf['EXT_KM'] = (valores / 1000).round(4)
            print(f"  [UNIDADE] '{nome_arquivo}': '{col}' está em METROS "
                  f"(razão {razao:.0f}x) -> convertido para km.")
        elif 0.9 < razao < 1.1:         # já em km
            gdf['EXT_KM'] = valores.round(4)
        else:
            print(f"  [QUALIDADE] '{nome_arquivo}': '{col}' não bate com a "
                  f"geometria (razão {razao:.2f}x). Extensão não exportada.")
            return gdf
        gdf = gdf.drop(columns=[col])
    except Exception as e:
        print(f"  Aviso: não normalizei a extensão de '{nome_arquivo}': {e}")
    return gdf


# Camadas oferecidas para download: (arquivo, padrão de busca, rótulo)
EXPORT_CAMADAS = [
    ('trechos',         'R*_TRECHOS.shp',          'Trechos'),
    ('pontos_sre',      'R*_PONTOS_SRE.shp',       'Pontos SRE'),
    ('regioes',         'R*_REGIÃO.shp',           'Regiões'),
    ('pontos_criticos', 'R*_Pontos_Criticos.shp',  'Pontos críticos'),
]


def exportar_dados(base_dir, camadas_dir):
    """
    Gera os arquivos de download em 'dados/': cada camada em KML (abre no
    Google Earth) e em Shapefile zipado (abre no ArcGIS/QGIS).

    As regiões vão juntas, com uma coluna REGIAO para filtrar. Os atributos
    saem LIMPOS (sem o \\r\\n que contamina os shapefiles de origem), então o
    download entrega um dado melhor que o original.
    """
    import zipfile
    import shutil
    import pandas as pd

    destino = base_dir / 'dados'
    destino.mkdir(exist_ok=True)
    tmp = destino / '_tmp'
    catalogo = []

    for arq, padrao, rotulo in EXPORT_CAMADAS:
        partes = []
        for f in sorted(camadas_dir.glob(padrao)):
            try:
                g = gpd.read_file(f)
                if g.crs and g.crs.to_string() != 'EPSG:4326':
                    g = g.to_crs(epsg=4326)
                elif not g.crs:
                    g.set_crs(epsg=4326, inplace=True)
                g = limpar_atributos(g)
                if 'TRECHOS' in f.stem.upper():
                    g = normalizar_extensao(g, f.stem)
                mm = re.match(r'^(R\d+)_', f.stem.upper())
                g['REGIAO'] = mm.group(1) if mm else '?'
                partes.append(g)
            except Exception as e:
                print(f"  Aviso: não exportei {f.name}: {e}")
        if not partes:
            continue

        # Só as colunas presentes em TODAS as regiões. Os shapefiles têm
        # schemas divergentes (ex.: 'GRUPO4' existe só no R1_REGIÃO). Juntar
        # tudo geraria colunas com nulo, e o driver KML falha ao escrever um
        # campo numérico nulo. Um dado para download também deve ter schema
        # uniforme entre as regiões.
        comuns = set(partes[0].columns)
        for p in partes[1:]:
            comuns &= set(p.columns)
        fora = set().union(*(set(p.columns) for p in partes)) - comuns
        if fora:
            print(f"  Aviso: '{rotulo}' - colunas fora do padrão, não exportadas: "
                  f"{', '.join(sorted(fora))}")
        ordem = [c for c in partes[0].columns if c in comuns]
        partes = [p[ordem] for p in partes]

        g = pd.concat(partes, ignore_index=True)

        try:
            # O driver KML do GDAL ANEXA a um arquivo existente em vez de
            # sobrescrever — o que faz a escrita falhar na 2ª feição. Por isso
            # apagamos antes. (Restos de uma escrita interrompida também
            # envenenariam a próxima execução.)
            kml = destino / f'{arq}.kml'
            if kml.exists():
                kml.unlink()
            g.to_file(kml, driver='KML')

            if tmp.exists():
                shutil.rmtree(tmp)
            tmp.mkdir()
            g.to_file(tmp / f'{arq}.shp')
            zp = destino / f'{arq}_shp.zip'
            if zp.exists():
                zp.unlink()
            with zipfile.ZipFile(zp, 'w', zipfile.ZIP_DEFLATED) as zf:
                for p in tmp.iterdir():
                    zf.write(p, p.name)
            shutil.rmtree(tmp)

            catalogo.append({
                'nome': rotulo,
                'kml': f'dados/{arq}.kml',
                'shp': f'dados/{arq}_shp.zip',
                'n': '{:,}'.format(len(g)).replace(',', '.'),
                'tam_kml': f'{kml.stat().st_size/1048576:.1f} MB',
                'tam_shp': f'{zp.stat().st_size/1048576:.1f} MB',
            })
            print(f"  Download: {rotulo} ({len(g)} feições) -> KML e SHP")
        except Exception as e:
            print(f"  Aviso: falhou ao exportar '{rotulo}': {e}")

    if tmp.exists():
        shutil.rmtree(tmp)
    return catalogo


def js_safe(valor):
    """Texto seguro para embutir dentro de uma string JavaScript."""
    if valor is None:
        return ''
    return (str(valor).replace('\\', ' ').replace('"', ' ')
            .replace('\r', ' ').replace('\n', ' ').strip())


class PainelControle(MacroElement):
    """
    Painel de controle unificado (estilo geoportal) com identidade RTA.
    Árvore de 3 níveis: Região > Trechos > Situação (PPS/LEN/...).
    """

    def __init__(self, basemaps, default_base, regioes, contexto,
                 filtros_situacao, subtitulo, logo=None, downloads=None,
                 filtros_criticos=None, filtros_inventario=None):
        super().__init__()
        self._name = 'PainelControle'
        self.basemaps = basemaps
        self.default_base = default_base
        self.regioes = regioes
        self.contexto = contexto
        self.filtros_situacao = filtros_situacao
        self.filtros_criticos = filtros_criticos or []
        self.filtros_inventario = filtros_inventario or []
        self.downloads = downloads or []
        self.subtitulo = subtitulo
        self.logo = logo
        self._template = Template(u"""
        {% macro html(this, kwargs) %}
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
            #gp-painel { font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                position: relative;
                width: 272px; max-height: 86vh; overflow-y: auto;
                background: rgba(255,255,255,0.97);
                -webkit-backdrop-filter: blur(16px) saturate(1.3); backdrop-filter: blur(16px) saturate(1.3);
                border: 1px solid rgba(255,255,255,0.6); border-radius: 16px;
                box-shadow: 0 2px 6px rgba(0,0,0,0.10), 0 16px 44px rgba(0,0,0,0.26);
                padding: 15px 16px; color: #1e293b; user-select: none;
                animation: gp-in 0.4s cubic-bezier(0.16,1,0.3,1); }
            @keyframes gp-in { from { opacity:0; transform:translateX(-10px);} to { opacity:1; transform:translateX(0);} }
            #gp-painel::-webkit-scrollbar { width: 6px; }
            #gp-painel::-webkit-scrollbar-thumb { background: rgba(0,0,0,0.15); border-radius: 3px; }

            #gp-painel .gp-header { display:flex; flex-direction:column; align-items:flex-start; gap:8px; padding-bottom:13px; margin-bottom:2px; border-bottom:1px solid rgba(0,0,0,0.08); }
            #gp-painel .gp-logo-img { width:154px; height:auto; display:block; }
            #gp-painel .gp-sub { display:flex; align-items:center; gap:6px; font-size:10.5px; font-weight:500; letter-spacing:0.02em; color:#64748b; }
            #gp-painel .gp-sub::before { content:''; width:6px; height:6px; border-radius:50%; background:#1E3A72; flex:0 0 auto; }

            #gp-painel .gp-sec { padding:13px 0; border-bottom:1px solid rgba(0,0,0,0.06); }
            #gp-painel .gp-sec:last-child { border-bottom:none; padding-bottom:2px; }
            #gp-painel .gp-sec-h { display:flex; align-items:center; gap:7px; font-size:10.5px; font-weight:700; letter-spacing:0.07em; text-transform:uppercase; color:#64748b; margin-bottom:9px; }
            #gp-painel .gp-sec-h svg { width:14px; height:14px; flex:0 0 auto; color:#1E3A72; }
            #gp-painel .gp-limpar { margin-left:auto; font-family:inherit; font-size:9px; font-weight:600;
                letter-spacing:0.02em; text-transform:none; color:#64748b; background:transparent;
                border:1px solid #e2e8f0; border-radius:20px; padding:2px 8px; cursor:pointer;
                transition:all .15s; }
            #gp-painel .gp-limpar:hover { color:#1E3A72; border-color:#1E3A72; background:rgba(30,58,114,0.06); }

            /* Busca */
            #gp-painel .gp-busca-wrap { position:relative; }
            #gp-painel #gp-busca { width:100%; box-sizing:border-box; height:32px; padding:0 26px 0 30px;
                border:1px solid #e2e8f0; border-radius:9px; background:#f8fafc;
                font-family:inherit; font-size:12px; color:#0f172a; outline:none; transition:border-color .15s, background .15s; }
            #gp-painel #gp-busca::placeholder { color:#94a3b8; }
            #gp-painel #gp-busca:focus { border-color:#1E3A72; background:#fff; box-shadow:0 0 0 3px rgba(30,58,114,0.10); }
            #gp-painel .gp-lupa { position:absolute; left:9px; top:50%; transform:translateY(-50%); width:14px; height:14px; color:#94a3b8; pointer-events:none; }
            #gp-painel #gp-limpar { position:absolute; right:6px; top:50%; transform:translateY(-50%); display:none;
                width:18px; height:18px; border:none; border-radius:50%; background:#cbd5e1; color:#fff;
                font-size:12px; line-height:1; cursor:pointer; padding:0; }
            #gp-painel #gp-limpar:hover { background:#94a3b8; }
            #gp-painel #gp-resultados { max-height:172px; overflow-y:auto; margin-top:6px; }
            #gp-painel #gp-resultados::-webkit-scrollbar { width:5px; }
            #gp-painel #gp-resultados::-webkit-scrollbar-thumb { background:rgba(0,0,0,0.12); border-radius:3px; }
            #gp-painel .gp-res { display:flex; align-items:center; gap:7px; padding:5px 7px; border-radius:6px; cursor:pointer; transition:background .12s; }
            #gp-painel .gp-res:hover { background:rgba(30,58,114,0.08); }
            #gp-painel .gp-res-txt { flex:1; min-width:0; display:flex; flex-direction:column; gap:1px; }
            #gp-painel .gp-res-sre { font-size:10.5px; font-weight:700; color:#1E3A72;
                font-family:ui-monospace,'SF Mono',Menlo,Consolas,monospace; }
            #gp-painel .gp-res-sub { font-size:9px; color:#94a3b8; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
            #gp-painel .gp-res-tag { font-size:8.5px; font-weight:700; color:#fff; background:var(--c);
                padding:2px 5px; border-radius:4px; letter-spacing:0.03em; flex:0 0 auto; }
            #gp-painel .gp-vazio { font-size:10.5px; color:#94a3b8; padding:8px 7px; text-align:center; }
            #gp-painel .gp-res-cab { display:flex; align-items:center; gap:8px; padding:3px 7px 5px;
                font-size:9.5px; font-weight:600; color:#94a3b8; border-bottom:1px solid #f1f5f9; margin-bottom:3px; }
            #gp-painel .gp-res-cab span { flex:1; }
            #gp-painel .gp-enquadrar { font-family:inherit; font-size:9px; font-weight:700; color:#1E3A72;
                background:rgba(30,58,114,0.09); border:none; border-radius:20px; padding:3px 9px;
                cursor:pointer; transition:background .15s; }
            #gp-painel .gp-enquadrar:hover { background:rgba(30,58,114,0.20); }

            /* Baixar dados */
            #gp-painel .gp-dl { display:flex; align-items:center; gap:7px; padding:5px 4px; margin:0 -4px;
                border-radius:6px; transition:background .12s; }
            #gp-painel .gp-dl:hover { background:rgba(15,23,42,0.04); }
            #gp-painel .gp-dl-txt { flex:1; min-width:0; display:flex; flex-direction:column; gap:1px; }
            #gp-painel .gp-dl-nome { font-size:11px; font-weight:600; color:#334155; }
            #gp-painel .gp-dl-n { font-size:8.5px; color:#94a3b8; font-variant-numeric:tabular-nums; }
            #gp-painel .gp-dl-bt { font-size:8.5px; font-weight:700; letter-spacing:0.04em; color:#1E3A72;
                background:rgba(30,58,114,0.09); border:none; border-radius:5px; padding:3px 7px;
                text-decoration:none; transition:background .15s; flex:0 0 auto; }
            #gp-painel .gp-dl-bt:hover { background:rgba(30,58,114,0.22); }
            #gp-painel .gp-dl-nota { font-size:9px; color:#94a3b8; line-height:1.45; padding-top:7px;
                margin-top:5px; border-top:1px solid #f1f5f9; }

            /* Chips de filtro (situação, pontos críticos, inventário) */
            #gp-painel .gp-filtro-tit { font-size:9px; font-weight:700; letter-spacing:0.04em;
                text-transform:uppercase; color:#94a3b8; margin:11px 0 5px; }
            #gp-painel .gp-filtro-tit span { font-weight:500; text-transform:none; letter-spacing:0; color:#cbd5e1; }
            #gp-painel .gp-chips { display:flex; flex-wrap:wrap; gap:4px; margin-top:0; }
            #gp-painel .gp-chip { display:inline-flex; align-items:center; gap:4px; padding:3px 7px;
                border:1px solid #e2e8f0; border-radius:20px; background:#fff; cursor:pointer;
                font-family:inherit; font-size:9.5px; font-weight:700; color:#cbd5e1;
                letter-spacing:0.04em; transition:all .15s; }
            #gp-painel .gp-chip::before { content:''; width:7px; height:7px; border-radius:50%; background:#e2e8f0; transition:background .15s; }
            #gp-painel .gp-chip.on { color:#334155; border-color:color-mix(in srgb, var(--c) 45%, #e2e8f0); background:color-mix(in srgb, var(--c) 8%, #fff); }
            #gp-painel .gp-chip.on::before { background:var(--c); }
            #gp-painel .gp-chip:hover { border-color:var(--c); }
            #gp-painel .gp-chip b { font-weight:600; opacity:0.65; font-variant-numeric:tabular-nums; }

            /* Mapas de fundo: cartões com ícone */
            #gp-painel .gp-bases { display:grid; grid-template-columns:repeat(3,1fr); gap:5px; }
            #gp-painel .gp-base-card { position:relative; cursor:pointer; }
            #gp-painel .gp-base-card input { position:absolute; opacity:0; width:0; height:0; }
            #gp-painel .gp-base-in { display:flex; flex-direction:column; align-items:center; gap:5px;
                padding:9px 3px; border:1px solid #e2e8f0; border-radius:10px; background:#f8fafc;
                transition:border-color .15s, background .15s, box-shadow .15s; }
            #gp-painel .gp-base-in svg { width:17px; height:17px; color:#94a3b8; transition:color .15s; }
            #gp-painel .gp-base-lbl { font-size:9.5px; font-weight:600; color:#64748b; letter-spacing:0.02em; }
            #gp-painel .gp-base-card:hover .gp-base-in { border-color:#cbd5e1; background:#fff; }
            #gp-painel .gp-base-card input:checked + .gp-base-in { border-color:#1E3A72;
                background:rgba(30,58,114,0.07); box-shadow:0 0 0 2px rgba(30,58,114,0.10); }
            #gp-painel .gp-base-card input:checked + .gp-base-in svg { color:#1E3A72; }
            #gp-painel .gp-base-card input:checked + .gp-base-in .gp-base-lbl { color:#1E3A72; font-weight:700; }
            #gp-painel .gp-base-card input:focus-visible + .gp-base-in { outline:2px solid #1E3A72; outline-offset:2px; }

            /* Nível 1: região. A linha é tingida com a cor da própria região
               (a mesma do polígono no mapa) e ganha uma barra à esquerda, para
               casar o nome com o que se vê no mapa sem precisar decorar. */
            #gp-painel .gp-reg { margin-bottom:2px; }
            #gp-painel .gp-reg-head { display:flex; align-items:center; gap:8px;
                padding:7px 6px 7px 10px; margin:1px -6px; border-radius:7px; cursor:pointer;
                background:color-mix(in srgb, var(--cr) 15%, transparent);
                box-shadow:inset 3px 0 0 var(--cr);
                transition:background .15s, box-shadow .15s; }
            #gp-painel .gp-reg-head:hover { background:color-mix(in srgb, var(--cr) 26%, transparent); }
            /* Região desligada: a tinta some junto com o resto */
            #gp-painel .gp-reg.desligada > .gp-reg-head { background:transparent;
                box-shadow:inset 3px 0 0 #cbd5e1; }
            #gp-painel .gp-chev { flex:0 0 auto; width:11px; color:#1E3A72; transition:transform .2s; }
            #gp-painel .gp-reg.open > .gp-reg-head .gp-chev { transform:rotate(90deg); }
            #gp-painel .gp-reg-nome { flex:1; font-size:13px; font-weight:600; color:#0f172a; }
            #gp-painel .gp-reg-body { max-height:0; overflow:hidden; transition:max-height .3s ease; padding-left:15px; }
            #gp-painel .gp-reg.open > .gp-reg-body { max-height:620px; }

            /* Nível 2: item (Formato / Trechos / Pontos SRE) */
            #gp-painel .gp-sub-item { display:flex; align-items:center; gap:9px; padding:5px 6px; margin:1px -6px; border-radius:7px; }
            #gp-painel .gp-sub-item:hover { background:rgba(15,23,42,0.04); }
            #gp-painel .gp-sub-lbl { flex:1; font-size:12.5px; color:#475569; }
            #gp-painel .gp-tre-head { display:flex; align-items:center; gap:9px; padding:5px 6px; margin:1px -6px; border-radius:7px; cursor:pointer; transition:background .15s; }
            #gp-painel .gp-tre-head:hover { background:rgba(15,23,42,0.04); }
            #gp-painel .gp-chev2 { flex:0 0 auto; width:9px; color:#1E3A72; transition:transform .2s; }
            #gp-painel .gp-tre.open > .gp-tre-head .gp-chev2 { transform:rotate(90deg); }
            #gp-painel .gp-tre-body { max-height:0; overflow:hidden; transition:max-height .25s ease; padding-left:14px; }
            #gp-painel .gp-tre.open > .gp-tre-body { max-height:440px; }

            /* Nível 3: situação */
            #gp-painel .gp-sit-head { display:flex; align-items:center; gap:8px; padding:3px 6px; margin:1px -6px; border-radius:6px; cursor:pointer; transition:background .15s; }
            #gp-painel .gp-sit-head:hover { background:rgba(15,23,42,0.04); }
            #gp-painel .gp-chev3 { flex:0 0 auto; width:8px; color:#1E3A72; opacity:0.75; transition:transform .2s, opacity .2s; }
            #gp-painel .gp-sit.open > .gp-sit-head .gp-chev3 { transform:rotate(90deg); opacity:1; }
            #gp-painel .gp-sit-txt { flex:1; min-width:0; display:flex; flex-direction:column; gap:1px; padding:1px 0; }
            #gp-painel .gp-sit-cod { font-size:10.5px; font-weight:700; color:#475569; letter-spacing:0.05em;
                font-family:ui-monospace,'SF Mono',Menlo,Consolas,monospace; }
            #gp-painel .gp-sit-desc { font-size:9.5px; line-height:1.25; color:#94a3b8; }

            /* Nível 4: lista de SREs (clique = voa até o trecho) */
            #gp-painel .gp-sre-body { max-height:0; overflow:hidden; transition:max-height .25s ease; }
            #gp-painel .gp-sit.open > .gp-sre-body { max-height:150px; overflow-y:auto; }
            #gp-painel .gp-sre-body::-webkit-scrollbar { width:5px; }
            #gp-painel .gp-sre-body::-webkit-scrollbar-thumb { background:rgba(0,0,0,0.12); border-radius:3px; }
            #gp-painel .gp-sre { display:flex; align-items:center; gap:6px; padding:3px 6px 3px 22px; margin:0 -6px;
                font-size:10.5px; font-family:ui-monospace,'SF Mono',Menlo,Consolas,monospace;
                color:#64748b; cursor:pointer; border-radius:5px; transition:background .12s, color .12s; }
            #gp-painel .gp-sre:hover { background:rgba(30,58,114,0.09); color:#1E3A72; }
            #gp-painel .gp-sre::before { content:''; width:4px; height:4px; border-radius:50%; background:currentColor; opacity:0.5; flex:0 0 auto; }

            /* MINIMALISTA: a linha é o controle. Não há interruptores — os
               <input> continuam existindo (toda a lógica de camadas depende
               deles), mas invisíveis. O estado se lê pela própria linha. */
            #gp-painel .gp-switch { display:none; }
            #gp-painel input[type="checkbox"] { position:absolute; opacity:0; width:0; height:0; pointer-events:none; }

            /* Cabeçalhos expandem ao clicar; as folhas não fazem nada no corpo
               (ligar/desligar é só pelo olho). */
            #gp-painel .gp-reg-head, #gp-painel .gp-tre-head, #gp-painel .gp-sit-head { cursor:pointer; }
            #gp-painel .off { opacity:0.45; }
            #gp-painel .off .gp-sub-lbl, #gp-painel .off .gp-sit-cod {
                text-decoration:line-through; text-decoration-color:#94a3b8; }
            #gp-painel .off .sw { filter:grayscale(1); }

            /* O olho é o ÚNICO botão de ligar/desligar — sempre visível
               (indispensável no celular, que não tem "passar o mouse"),
               destacando no hover e quando a camada está desligada. */
            #gp-painel .gp-olho { width:15px; height:15px; flex:0 0 auto; color:#94a3b8;
                opacity:0.5; cursor:pointer; transition:opacity .12s, color .12s; padding:2px; margin:-2px; box-sizing:content-box; }
            #gp-painel .gp-olho:hover { opacity:1; color:#1E3A72; }
            #gp-painel .gp-reg-head:hover .gp-olho, #gp-painel .gp-sub-item:hover .gp-olho,
            #gp-painel .gp-tre-head:hover .gp-olho, #gp-painel .gp-sit-head:hover .gp-olho { opacity:0.9; }
            #gp-painel .off .gp-olho, #gp-painel .gp-reg.desligada .gp-olho { opacity:1; color:#64748b; }
            #gp-painel .gp-olho .o-off { display:none; }
            #gp-painel .off .gp-olho .o-on, #gp-painel .gp-reg.desligada > .gp-reg-head .gp-olho .o-on { display:none; }
            #gp-painel .off .gp-olho .o-off, #gp-painel .gp-reg.desligada > .gp-reg-head .gp-olho .o-off { display:block; }

            /* Contadores em texto puro (a pílula cinza saiu com o redesenho),
               mas em tom escuro o bastante para leitura rápida em tela clara
               ou por quem tem baixa visão. */
            #gp-painel .gp-cnt { font-variant-numeric:tabular-nums; font-weight:600; color:#334155;
                font-size:10px; background:transparent; padding:0; border-radius:0; }
            #gp-painel .gp-cnt.xs { font-size:9.5px; font-weight:600; color:#475569; }

            /* Região desligada: todo o bloco apaga, deixando claro que as
               subcamadas foram ocultadas em lote. */
            #gp-painel .gp-reg.desligada > .gp-reg-body { opacity:0.4; }
            #gp-painel .gp-reg.desligada > .gp-reg-head .gp-reg-nome { color:#94a3b8; }

            /* Símbolos da legenda. Propositalmente NÃO parecem controles:
               o polígono é uma forma irregular (e não um quadrado, que
               lembraria um checkbox ao lado da chave liga/desliga). */
            #gp-painel .sw { flex:0 0 auto; display:inline-block; }
            #gp-painel .sw.ponto, #gp-painel .sw.crit { width:12px; height:12px; border-radius:50%; background:var(--c); border:2px solid #fff; box-shadow:0 0 0 1px rgba(0,0,0,0.18); }
            #gp-painel .sw.crit { box-shadow:0 0 0 1px rgba(0,0,0,0.12), 0 0 0 3px color-mix(in srgb, var(--c) 22%, transparent); }
            #gp-painel .sw.no { width:8px; height:8px; border-radius:50%; background:#fff;
                border:2px solid var(--c); box-shadow:0 0 0 1px rgba(0,0,0,0.10); }
            #gp-painel .sw.linha { width:20px; height:5px; border-radius:3px; background:var(--c); }
            #gp-painel .sw.linha-xs { width:16px; height:4px; border-radius:2px; background:var(--c); }
            #gp-painel .sw.poli { width:16px; height:14px; background:var(--c); opacity:0.5;
                clip-path:polygon(14% 6%, 76% 0%, 100% 52%, 64% 100%, 20% 90%, 0% 38%); }
            #gp-painel .sw.contorno { width:16px; height:14px; background:var(--c); opacity:0.65;
                clip-path:polygon(14% 6%, 76% 0%, 100% 52%, 64% 100%, 20% 90%, 0% 38%,
                                  6% 42%, 24% 82%, 62% 92%, 92% 52%, 72% 8%, 18% 14%); }
            #gp-painel .sw.multi { width:20px; height:5px; border-radius:3px;
                background:linear-gradient(90deg,#047857 0 25%,#2563EB 25% 50%,#F59E0B 50% 75%,#B45309 75% 100%); }
            /* Símbolo do grupo Inventário: quadradinho com um "pino" */
            #gp-painel .sw.inv { position:relative; width:13px; height:13px; border-radius:3px;
                border:1.5px solid var(--c); background:color-mix(in srgb, var(--c) 15%, transparent); }
            #gp-painel .sw.inv::after { content:''; position:absolute; inset:3px; border-radius:50%; background:var(--c); }

            /* Acessibilidade: foco visível por teclado */
            #gp-painel input:focus-visible + .gp-dot,
            #gp-painel input:focus-visible + .gp-slider { outline:2px solid #1E3A72; outline-offset:2px; }

            /* --- Brilho dos pontos críticos no mapa ---------------------
               Halo constante (drop-shadow) para o ponto não sumir sobre o
               satélite, mais uma pulsação de ~2,4 s que toca sozinha quando a
               camada é ativada: o Leaflet recria o <path>, e o CSS reinicia a
               animação. Depois assenta no halo fixo, sem ficar piscando. */
            .pc-mk { transition: filter .2s ease; }
            .pc-c1 { filter: drop-shadow(0 0 3px rgba(220,38,38,.95)); animation: pc-surge1 2.4s ease-out 1; }
            .pc-c2 { filter: drop-shadow(0 0 3px rgba(245,158,11,.95)); animation: pc-surge2 2.4s ease-out 1; }
            .pc-c3 { filter: drop-shadow(0 0 3px rgba(5,150,105,.95));  animation: pc-surge3 2.4s ease-out 1; }
            .pc-c0 { filter: drop-shadow(0 0 2px rgba(148,163,184,.9)); }
            .pc-mk:hover { filter: drop-shadow(0 0 7px rgba(255,255,255,.95)); cursor:pointer; }

            @keyframes pc-surge1 {
                0%   { filter: drop-shadow(0 0 0 rgba(220,38,38,0)); }
                18%  { filter: drop-shadow(0 0 11px rgba(220,38,38,1)) drop-shadow(0 0 18px rgba(220,38,38,.75)); }
                55%  { filter: drop-shadow(0 0 4px rgba(220,38,38,.95)); }
                75%  { filter: drop-shadow(0 0 9px rgba(220,38,38,1)); }
                100% { filter: drop-shadow(0 0 3px rgba(220,38,38,.95)); }
            }
            @keyframes pc-surge2 {
                0%   { filter: drop-shadow(0 0 0 rgba(245,158,11,0)); }
                18%  { filter: drop-shadow(0 0 11px rgba(245,158,11,1)) drop-shadow(0 0 18px rgba(245,158,11,.75)); }
                55%  { filter: drop-shadow(0 0 4px rgba(245,158,11,.95)); }
                75%  { filter: drop-shadow(0 0 9px rgba(245,158,11,1)); }
                100% { filter: drop-shadow(0 0 3px rgba(245,158,11,.95)); }
            }
            @keyframes pc-surge3 {
                0%   { filter: drop-shadow(0 0 0 rgba(5,150,105,0)); }
                18%  { filter: drop-shadow(0 0 11px rgba(5,150,105,1)) drop-shadow(0 0 18px rgba(5,150,105,.75)); }
                55%  { filter: drop-shadow(0 0 4px rgba(5,150,105,.95)); }
                75%  { filter: drop-shadow(0 0 9px rgba(5,150,105,1)); }
                100% { filter: drop-shadow(0 0 3px rgba(5,150,105,.95)); }
            }
            /* Respeita quem pediu menos animação no sistema */
            @media (prefers-reduced-motion: reduce) {
                .pc-c1, .pc-c2, .pc-c3 { animation: none; }
            }

            /* O poligono da regiao e so pano de fundo: os cliques atravessam
               ele e chegam nos trechos e pontos que estao por cima. Sem isto,
               era preciso desligar a regiao para conseguir clicar no resto. */
            .gp-regiao { pointer-events: none !important; }

            /* Ficha do ponto crítico (popup) */
            .pc { font-family:'Inter',-apple-system,'Segoe UI',Roboto,sans-serif; color:#1e293b; min-width:236px; }
            .pc-top { display:flex; align-items:center; gap:8px; padding-bottom:7px; border-bottom:1px solid #e2e8f0; }
            .pc-num { flex:1; font-size:13px; font-weight:700; color:#0f172a; }
            .pc-tag { font-size:8.5px; font-weight:700; color:#fff; padding:3px 7px; border-radius:20px; letter-spacing:0.03em; }
            .pc-sub { font-size:10.5px; color:#64748b; padding:6px 0 2px; line-height:1.35; }
            .pc-hist { margin:6px 0 0; border-top:1px solid #f1f5f9; padding-top:5px; }
            .pc-linha { display:flex; align-items:center; gap:8px; padding:2.5px 0; font-size:11px; }
            .pc-mes { width:26px; flex:0 0 auto; font-weight:700; color:#94a3b8; font-size:9.5px;
                text-transform:uppercase; letter-spacing:0.04em; }
            .pc-st { flex:1; font-weight:600; font-size:10.5px; }
            .pc-foto { font-size:9.5px; font-weight:600; color:#1E3A72; text-decoration:none;
                background:rgba(30,58,114,0.08); padding:2px 7px; border-radius:20px; white-space:nowrap; }
            .pc-foto:hover { background:rgba(30,58,114,0.16); text-decoration:underline; }
            .pc-nofoto { font-size:9.5px; color:#cbd5e1; padding:2px 7px; }
            .pc-final { margin-top:7px; padding-top:7px; border-top:1px solid #f1f5f9;
                font-size:10.5px; line-height:1.45; color:#475569; max-height:130px; overflow-y:auto; }
            .leaflet-popup-content { margin:11px 13px; }
            .leaflet-popup-content-wrapper { border-radius:10px; }

            /* ---------- Celular: painel vira uma "bottom sheet" ----------
               No computador nada muda. No celular o painel comeca fechado (o
               mapa aparece inteiro); um botao "Camadas" no rodape sobe um
               painel que ocupa so a metade de baixo, deixando o mapa visivel
               em cima. Pensado para quem esta em campo. */
            .gp-abrir { display:none; align-items:center; gap:8px; padding:11px 18px;
                border:none; border-radius:26px; background:#1E3A72; color:#fff; cursor:pointer;
                font-family:'Inter',-apple-system,sans-serif; font-size:14px; font-weight:600;
                box-shadow:0 4px 16px rgba(0,0,0,0.35); }
            .gp-abrir svg { width:20px; height:20px; flex:0 0 auto; }
            #gp-painel .gp-fechar { display:none; position:absolute; top:11px; right:12px;
                width:32px; height:32px; padding:0; border:none; border-radius:9px;
                background:rgba(15,23,42,0.06); color:#475569; cursor:pointer;
                align-items:center; justify-content:center; z-index:3; }
            #gp-painel .gp-fechar svg { width:17px; height:17px; }
            #gp-painel .gp-fechar:hover { background:rgba(15,23,42,0.12); }
            #gp-painel .gp-alca { display:none; position:absolute; top:9px; left:50%;
                transform:translateX(-50%); width:42px; height:4px; border-radius:2px;
                background:#cbd5e1; }

            @media (max-width: 640px) {
                #gp-painel {
                    position:fixed !important; left:0 !important; right:0 !important;
                    bottom:0 !important; top:auto !important; width:auto !important;
                    max-width:none; margin:0 !important;
                    max-height:60vh; border-radius:18px 18px 0 0; padding-top:22px;
                    box-shadow:0 -6px 30px rgba(0,0,0,0.30);
                    display:none; animation:gp-subir 0.28s cubic-bezier(0.16,1,0.3,1); }
                #gp-painel.aberto { display:block; }
                #gp-painel .gp-alca { display:block; }
                #gp-painel .gp-fechar { display:flex; }
                .gp-abrir { display:flex; }
                .gp-abrir.escondido { display:none; }
            }
            @keyframes gp-subir { from { transform:translateY(100%); } to { transform:translateY(0); } }
        </style>

        <div id="gp-painel">
            <div class="gp-alca"></div>
            <button type="button" class="gp-fechar" aria-label="Fechar painel">
                <svg viewBox="0 0 24 24" fill="none"><path d="M6 6l12 12M18 6L6 18" stroke="currentColor" stroke-width="2.2" stroke-linecap="round"/></svg>
            </button>
            <div class="gp-header">
                {% if this.logo %}
                <img class="gp-logo-img" src="{{ this.logo }}" alt="RTA Engenheiros Consultores">
                {% endif %}
                <div class="gp-sub">{{ this.subtitulo }}</div>
            </div>

            <div class="gp-sec">
                <div class="gp-sec-h">
                    <svg viewBox="0 0 24 24" fill="none"><path d="M12 2l9 5-9 5-9-5 9-5zM3 12l9 5 9-5M3 17l9 5 9-5" stroke="currentColor" stroke-width="1.7" stroke-linejoin="round"/></svg>
                    Mapas de Fundo
                </div>
                <div class="gp-bases">
                    {% for b in this.basemaps %}
                    <label class="gp-base-card" title="{{ b.nome }}">
                        <input type="radio" name="gp-base" value="{{ b.layer.get_name() }}" {% if b.layer.get_name() == this.default_base.get_name() %}checked{% endif %}>
                        <span class="gp-base-in">
                            {% if b.icone == 'mapa' %}
                            <svg viewBox="0 0 24 24" fill="none"><path d="M9 3L3 5.5v15L9 18l6 2.5 6-2.5v-15L15 5.5 9 3z" stroke="currentColor" stroke-width="1.7" stroke-linejoin="round"/><path d="M9 3v15M15 5.5v15" stroke="currentColor" stroke-width="1.7"/></svg>
                            {% elif b.icone == 'satelite' %}
                            <svg viewBox="0 0 24 24" fill="none"><circle cx="12" cy="12" r="9" stroke="currentColor" stroke-width="1.7"/><path d="M3 12h18M12 3c2.5 2.6 2.5 15.4 0 18M12 3c-2.5 2.6-2.5 15.4 0 18" stroke="currentColor" stroke-width="1.5"/></svg>
                            {% else %}
                            <svg viewBox="0 0 24 24" fill="none"><path d="M20.5 14.5A8.5 8.5 0 019.5 3.5a8.5 8.5 0 1011 11z" stroke="currentColor" stroke-width="1.7" stroke-linejoin="round"/></svg>
                            {% endif %}
                            <span class="gp-base-lbl">{{ b.nome }}</span>
                        </span>
                    </label>
                    {% endfor %}
                </div>
            </div>

            <div class="gp-sec">
                <div class="gp-sec-h">
                    <svg viewBox="0 0 24 24" fill="none"><path d="M3 6l9-4 9 4-9 4-9-4zM3 12l9 4 9-4M3 18l9 4 9-4" stroke="currentColor" stroke-width="1.6" stroke-linejoin="round"/></svg>
                    Camadas por Região
                    <button type="button" id="gp-limpar-regioes" class="gp-limpar">Desmarcar todas</button>
                </div>
                {% for reg in this.regioes %}
                <div class="gp-reg {% if loop.first %}open{% endif %}" data-reg="{{ reg.id }}" style="--cr: {{ reg.cor }}">
                    <div class="gp-reg-head">
                        <svg class="gp-chev" viewBox="0 0 24 24" fill="none"><path d="M9 6l6 6-6 6" stroke="currentColor" stroke-width="2.4" stroke-linecap="round" stroke-linejoin="round"/></svg>
                        <span class="gp-reg-nome">{{ reg.nome }}</span>
                        <span class="gp-olho" data-olho-reg="{{ reg.id }}" title="Mostrar/ocultar a região">
                            <svg class="o-on" viewBox="0 0 24 24" fill="none"><path d="M2 12s3.6-7 10-7 10 7 10 7-3.6 7-10 7S2 12 2 12z" stroke="currentColor" stroke-width="1.9"/><circle cx="12" cy="12" r="2.6" stroke="currentColor" stroke-width="1.9"/></svg>
                            <svg class="o-off" viewBox="0 0 24 24" fill="none"><path d="M3 3l18 18M10.6 10.7a2.6 2.6 0 003.7 3.7M9.4 5.3A9.5 9.5 0 0112 5c6.4 0 10 7 10 7a17 17 0 01-3.2 4M6.2 6.2A17 17 0 002 12s3.6 7 10 7c1.3 0 2.4-.2 3.5-.6" stroke="currentColor" stroke-width="1.9" stroke-linecap="round"/></svg>
                        </span>
                        <input type="checkbox" data-regiao="{{ reg.id }}" checked>
                    </div>
                    <div class="gp-reg-body">
                        {% for it in reg.itens %}
                        {% if it.situacoes %}
                        <div class="gp-tre" data-grp="{{ it.grupo }}">
                            <div class="gp-tre-head">
                                <svg class="gp-chev2" viewBox="0 0 24 24" fill="none"><path d="M9 6l6 6-6 6" stroke="currentColor" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/></svg>
                                <span class="sw {{ it.forma }}" style="--c: {{ it.cor }}"></span>
                                <span class="gp-sub-lbl">{{ it.nome }}</span>
                                <span class="gp-cnt">{{ it.count }}</span>
                                <span class="gp-olho">
                                    <svg class="o-on" viewBox="0 0 24 24" fill="none"><path d="M2 12s3.6-7 10-7 10 7 10 7-3.6 7-10 7S2 12 2 12z" stroke="currentColor" stroke-width="1.9"/><circle cx="12" cy="12" r="2.6" stroke="currentColor" stroke-width="1.9"/></svg>
                                    <svg class="o-off" viewBox="0 0 24 24" fill="none"><path d="M3 3l18 18M10.6 10.7a2.6 2.6 0 003.7 3.7M9.4 5.3A9.5 9.5 0 0112 5c6.4 0 10 7 10 7a17 17 0 01-3.2 4M6.2 6.2A17 17 0 002 12s3.6 7 10 7c1.3 0 2.4-.2 3.5-.6" stroke="currentColor" stroke-width="1.9" stroke-linecap="round"/></svg>
                                </span>
                                <input type="checkbox" data-regiao="{{ reg.id }}" data-grupo="{{ it.grupo }}" checked>
                            </div>
                            <div class="gp-tre-body">
                                {% for s in it.situacoes %}
                                <div class="gp-sit">
                                    <div class="gp-sit-head">
                                        <svg class="gp-chev3" viewBox="0 0 24 24" fill="none"><path d="M9 6l6 6-6 6" stroke="currentColor" stroke-width="3.4" stroke-linecap="round" stroke-linejoin="round"/></svg>
                                        <input type="checkbox" data-regiao="{{ reg.id }}" data-grupo="{{ it.grupo }}" data-sit="{{ s.codigo }}" data-camada="{{ s.layer.get_name() }}" checked>
                                        <span class="sw linha-xs" style="--c: {{ s.cor }}"></span>
                                        <span class="gp-sit-txt">
                                            <span class="gp-sit-cod">{{ s.codigo }}</span>
                                            {% if s.desc %}<span class="gp-sit-desc">{{ s.desc }}</span>{% endif %}
                                        </span>
                                        <span class="gp-cnt xs">{{ s.count }}</span>
                                        <span class="gp-olho">
                                            <svg class="o-on" viewBox="0 0 24 24" fill="none"><path d="M2 12s3.6-7 10-7 10 7 10 7-3.6 7-10 7S2 12 2 12z" stroke="currentColor" stroke-width="1.9"/><circle cx="12" cy="12" r="2.6" stroke="currentColor" stroke-width="1.9"/></svg>
                                            <svg class="o-off" viewBox="0 0 24 24" fill="none"><path d="M3 3l18 18M10.6 10.7a2.6 2.6 0 003.7 3.7M9.4 5.3A9.5 9.5 0 0112 5c6.4 0 10 7 10 7a17 17 0 01-3.2 4M6.2 6.2A17 17 0 002 12s3.6 7 10 7c1.3 0 2.4-.2 3.5-.6" stroke="currentColor" stroke-width="1.9" stroke-linecap="round"/></svg>
                                        </span>
                                    </div>
                                    <div class="gp-sre-body">
                                        {% for sr in s.sres %}
                                        <div class="gp-sre" data-b="{{ sr.b }}" data-sre="{{ sr.sre }}" data-fg="{{ s.layer.get_name() }}" title="Ir para {{ sr.sre }}">{{ sr.sre }}</div>
                                        {% endfor %}
                                    </div>
                                </div>
                                {% endfor %}
                            </div>
                        </div>
                        {% elif it.subitens %}
                        <div class="gp-tre" data-grp="{{ it.grupo }}">
                            <div class="gp-tre-head">
                                <svg class="gp-chev2" viewBox="0 0 24 24" fill="none"><path d="M9 6l6 6-6 6" stroke="currentColor" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/></svg>
                                <span class="sw {{ it.forma }}" style="--c: {{ it.cor }}"></span>
                                <span class="gp-sub-lbl">{{ it.nome }}</span>
                                <span class="gp-cnt">{{ it.count }}</span>
                                <span class="gp-olho">
                                    <svg class="o-on" viewBox="0 0 24 24" fill="none"><path d="M2 12s3.6-7 10-7 10 7 10 7-3.6 7-10 7S2 12 2 12z" stroke="currentColor" stroke-width="1.9"/><circle cx="12" cy="12" r="2.6" stroke="currentColor" stroke-width="1.9"/></svg>
                                    <svg class="o-off" viewBox="0 0 24 24" fill="none"><path d="M3 3l18 18M10.6 10.7a2.6 2.6 0 003.7 3.7M9.4 5.3A9.5 9.5 0 0112 5c6.4 0 10 7 10 7a17 17 0 01-3.2 4M6.2 6.2A17 17 0 002 12s3.6 7 10 7c1.3 0 2.4-.2 3.5-.6" stroke="currentColor" stroke-width="1.9" stroke-linecap="round"/></svg>
                                </span>
                                <input type="checkbox" data-regiao="{{ reg.id }}" data-grupo="{{ it.grupo }}">
                            </div>
                            <div class="gp-tre-body">
                                {% for sc in it.subitens %}
                                <div class="gp-sub-item">
                                    <input type="checkbox" data-regiao="{{ reg.id }}" data-grupo="{{ it.grupo }}" data-inv="{{ sc.chave }}" data-camada="{{ sc.layer.get_name() }}">
                                    <span class="sw {{ sc.forma }}" style="--c: {{ sc.cor }}"></span>
                                    <span class="gp-sub-lbl">{{ sc.nome }}</span>
                                    <span class="gp-cnt">{{ sc.count }}</span>
                                    <span class="gp-olho">
                                        <svg class="o-on" viewBox="0 0 24 24" fill="none"><path d="M2 12s3.6-7 10-7 10 7 10 7-3.6 7-10 7S2 12 2 12z" stroke="currentColor" stroke-width="1.9"/><circle cx="12" cy="12" r="2.6" stroke="currentColor" stroke-width="1.9"/></svg>
                                        <svg class="o-off" viewBox="0 0 24 24" fill="none"><path d="M3 3l18 18M10.6 10.7a2.6 2.6 0 003.7 3.7M9.4 5.3A9.5 9.5 0 0112 5c6.4 0 10 7 10 7a17 17 0 01-3.2 4M6.2 6.2A17 17 0 002 12s3.6 7 10 7c1.3 0 2.4-.2 3.5-.6" stroke="currentColor" stroke-width="1.9" stroke-linecap="round"/></svg>
                                    </span>
                                </div>
                                {% endfor %}
                            </div>
                        </div>
                        {% else %}
                        <div class="gp-sub-item">
                            <input type="checkbox" data-regiao="{{ reg.id }}" data-camada="{{ it.layer.get_name() }}" checked>
                            <span class="sw {{ it.forma }}" style="--c: {{ it.cor }}"></span>
                            <span class="gp-sub-lbl">{{ it.nome }}</span>
                            <span class="gp-cnt">{{ it.count }}</span>
                            <span class="gp-olho">
                                <svg class="o-on" viewBox="0 0 24 24" fill="none"><path d="M2 12s3.6-7 10-7 10 7 10 7-3.6 7-10 7S2 12 2 12z" stroke="currentColor" stroke-width="1.9"/><circle cx="12" cy="12" r="2.6" stroke="currentColor" stroke-width="1.9"/></svg>
                                <svg class="o-off" viewBox="0 0 24 24" fill="none"><path d="M3 3l18 18M10.6 10.7a2.6 2.6 0 003.7 3.7M9.4 5.3A9.5 9.5 0 0112 5c6.4 0 10 7 10 7a17 17 0 01-3.2 4M6.2 6.2A17 17 0 002 12s3.6 7 10 7c1.3 0 2.4-.2 3.5-.6" stroke="currentColor" stroke-width="1.9" stroke-linecap="round"/></svg>
                            </span>
                        </div>
                        {% endif %}
                        {% endfor %}
                    </div>
                </div>
                {% endfor %}
            </div>

            {% if this.contexto %}
            <div class="gp-sec">
                <div class="gp-sec-h">
                    <svg viewBox="0 0 24 24" fill="none"><path d="M12 21s7-5.686 7-11a7 7 0 10-14 0c0 5.314 7 11 7 11z" stroke="currentColor" stroke-width="1.6"/><circle cx="12" cy="10" r="2.4" stroke="currentColor" stroke-width="1.6"/></svg>
                    Camadas de Contexto
                </div>
                {% for c in this.contexto %}
                <div class="gp-sub-item">
                    <input type="checkbox" data-camada="{{ c.layer.get_name() }}" {% if c.ativo %}checked{% endif %}>
                    <span class="sw {{ c.forma }}" style="--c: {{ c.cor }}"></span>
                    <span class="gp-sub-lbl">{{ c.nome }}</span>
                    <span class="gp-cnt">{{ c.count }}</span>
                    <span class="gp-olho">
                        <svg class="o-on" viewBox="0 0 24 24" fill="none"><path d="M2 12s3.6-7 10-7 10 7 10 7-3.6 7-10 7S2 12 2 12z" stroke="currentColor" stroke-width="1.9"/><circle cx="12" cy="12" r="2.6" stroke="currentColor" stroke-width="1.9"/></svg>
                        <svg class="o-off" viewBox="0 0 24 24" fill="none"><path d="M3 3l18 18M10.6 10.7a2.6 2.6 0 003.7 3.7M9.4 5.3A9.5 9.5 0 0112 5c6.4 0 10 7 10 7a17 17 0 01-3.2 4M6.2 6.2A17 17 0 002 12s3.6 7 10 7c1.3 0 2.4-.2 3.5-.6" stroke="currentColor" stroke-width="1.9" stroke-linecap="round"/></svg>
                    </span>
                </div>
                {% endfor %}
            </div>
            {% endif %}

            <div class="gp-sec">
                <div class="gp-sec-h">
                    <svg viewBox="0 0 24 24" fill="none"><circle cx="11" cy="11" r="7" stroke="currentColor" stroke-width="1.8"/><path d="M20 20l-3.5-3.5" stroke="currentColor" stroke-width="1.8" stroke-linecap="round"/></svg>
                    Busca &amp; Filtros
                </div>
                <div class="gp-busca-wrap">
                    <svg class="gp-lupa" viewBox="0 0 24 24" fill="none"><circle cx="11" cy="11" r="7" stroke="currentColor" stroke-width="2"/><path d="M20 20l-3.5-3.5" stroke="currentColor" stroke-width="2" stroke-linecap="round"/></svg>
                    <input id="gp-busca" type="text" placeholder="SRE, rodovia ou cidade…" autocomplete="off" aria-label="Buscar SRE, rodovia ou cidade">
                    <button id="gp-limpar" type="button" aria-label="Limpar busca">&times;</button>
                </div>
                <div id="gp-resultados"></div>
                <div class="gp-filtro-tit">Situação dos trechos</div>
                <div class="gp-chips">
                    {% for f in this.filtros_situacao %}
                    <button type="button" class="gp-chip on" data-sit="{{ f.codigo }}" style="--c: {{ f.cor }}" title="{{ f.desc }}">{{ f.codigo }} <b>{{ f.count }}</b></button>
                    {% endfor %}
                </div>

                {% if this.filtros_criticos %}
                <div class="gp-filtro-tit">Pontos críticos</div>
                <div class="gp-chips">
                    {% for f in this.filtros_criticos %}
                    <button type="button" class="gp-chip on" data-sit="{{ f.codigo }}" style="--c: {{ f.cor }}" title="{{ f.desc }}">{{ f.codigo }} <b>{{ f.count }}</b></button>
                    {% endfor %}
                </div>
                {% endif %}

                {% if this.filtros_inventario %}
                <div class="gp-filtro-tit">Inventário <span>— liga em todas as regiões</span></div>
                <div class="gp-chips">
                    {% for f in this.filtros_inventario %}
                    <button type="button" class="gp-chip" data-inv="{{ f.codigo }}" style="--c: {{ f.cor }}" title="{{ f.desc }}">{{ f.desc }} <b>{{ f.count }}</b></button>
                    {% endfor %}
                </div>
                {% endif %}
            </div>

            {% if this.downloads %}
            <div class="gp-sec">
                <div class="gp-sec-h">
                    <svg viewBox="0 0 24 24" fill="none"><path d="M12 3v12m0 0l-4.5-4.5M12 15l4.5-4.5M4 17v2a2 2 0 002 2h12a2 2 0 002-2v-2" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/></svg>
                    Baixar Dados
                </div>
                {% for d in this.downloads %}
                <div class="gp-dl">
                    <span class="gp-dl-txt">
                        <span class="gp-dl-nome">{{ d.nome }}</span>
                        <span class="gp-dl-n">{{ d.n }} feições</span>
                    </span>
                    <a class="gp-dl-bt" href="{{ d.kml }}" download title="KML · {{ d.tam_kml }} · abre no Google Earth">KML</a>
                    <a class="gp-dl-bt" href="{{ d.shp }}" download title="Shapefile zipado · {{ d.tam_shp }} · abre no ArcGIS/QGIS">SHP</a>
                </div>
                {% endfor %}
                <div class="gp-dl-nota">
                    KML abre no Google Earth · SHP no ArcGIS e QGIS.<br>
                    Todas as regiões juntas, com a coluna <b>REGIAO</b>. EPSG:4326.
                </div>
            </div>
            {% endif %}
        </div>
        {% endmacro %}

        {% macro script(this, kwargs) %}
            (function() {
                var map = {{ this._parent.get_name() }};

                var Painel = L.Control.extend({
                    options: { position: 'topleft' },
                    onAdd: function() {
                        var d = document.getElementById('gp-painel');
                        L.DomEvent.disableClickPropagation(d);
                        L.DomEvent.disableScrollPropagation(d);
                        return d;
                    }
                });
                new Painel().addTo(map);
                L.control.zoom({ position: 'topright' }).addTo(map);

                // --- Gaveta no celular: botão flutuante para abrir o painel ---
                var painelDiv = document.getElementById('gp-painel');
                var BotaoAbrir = L.Control.extend({
                    options: { position: 'bottomleft' },
                    onAdd: function() {
                        var b = L.DomUtil.create('button', 'gp-abrir');
                        b.type = 'button';
                        b.setAttribute('aria-label', 'Abrir camadas e filtros');
                        b.innerHTML = '<svg viewBox="0 0 24 24" fill="none"><path d="M3 6l9-4 9 4-9 4-9-4zM3 12l9 4 9-4M3 18l9 4 9-4" stroke="currentColor" stroke-width="1.7" stroke-linejoin="round"/></svg><span>Camadas</span>';
                        L.DomEvent.disableClickPropagation(b);
                        L.DomEvent.on(b, 'click', function() {
                            painelDiv.classList.add('aberto');
                            b.classList.add('escondido');
                        });
                        return b;
                    }
                });
                new BotaoAbrir().addTo(map);

                painelDiv.querySelector('.gp-fechar').addEventListener('click', function() {
                    painelDiv.classList.remove('aberto');
                    var ab = document.querySelector('.gp-abrir');
                    if (ab) ab.classList.remove('escondido');
                });

                // --- MINIMALISTA: a linha é o controle ---------------------
                // Regra: a seta expande/recolhe; o resto da linha liga/desliga.
                function alternar(inp) {
                    if (!inp) return;
                    inp.checked = !inp.checked;
                    inp.dispatchEvent(new Event('change', { bubbles: true }));
                }
                function pintarLinha(inp) {
                    var linha = inp.closest('.gp-sub-item, .gp-tre-head, .gp-sit-head');
                    if (linha) linha.classList.toggle('off', !inp.checked);
                }
                function pintarTudo() {
                    document.querySelectorAll('#gp-painel .gp-sub-item input, #gp-painel .gp-tre-head input, #gp-painel .gp-sit-head input')
                        .forEach(pintarLinha);
                }

                // Região: clicar na linha expande; o olho liga/desliga a região
                document.querySelectorAll('#gp-painel .gp-reg-head').forEach(function(h){
                    h.addEventListener('click', function(){ h.parentNode.classList.toggle('open'); });
                });
                document.querySelectorAll('#gp-painel .gp-olho[data-olho-reg]').forEach(function(o){
                    o.addEventListener('click', function(e){
                        e.stopPropagation();
                        alternar(o.parentNode.querySelector('input[data-regiao]'));
                    });
                });

                // Setas expandem (sem alternar a camada)
                document.querySelectorAll('#gp-painel .gp-chev2, #gp-painel .gp-chev3').forEach(function(ch){
                    ch.addEventListener('click', function(e){
                        e.stopPropagation();
                        ch.parentNode.parentNode.classList.toggle('open');
                    });
                });

                // Cabeçalhos de grupo/situação: clicar na linha EXPANDE.
                // (Ligar/desligar é só pelo olho, a pedido — a linha não desliga.)
                document.querySelectorAll('#gp-painel .gp-tre-head, #gp-painel .gp-sit-head')
                    .forEach(function(h){
                        h.addEventListener('click', function(){ h.parentNode.classList.toggle('open'); });
                    });

                // O olho é o ÚNICO controle de ligar/desligar (o da região já
                // tem seu handler acima; este cobre todo o resto).
                document.querySelectorAll('#gp-painel .gp-olho').forEach(function(o){
                    if (o.hasAttribute('data-olho-reg')) return;
                    o.addEventListener('click', function(e){
                        e.stopPropagation();
                        var linha = o.closest('.gp-sub-item, .gp-tre-head, .gp-sit-head');
                        if (linha) alternar(linha.querySelector('input[type="checkbox"]'));
                    });
                });

                // --- Mapas de fundo ---
                var bases = { {% for b in this.basemaps %}"{{ b.layer.get_name() }}": {{ b.layer.get_name() }}{% if not loop.last %}, {% endif %}{% endfor %} };
                var padrao = {{ this.default_base.get_name() }};
                Object.keys(bases).forEach(function(k){ if (map.hasLayer(bases[k])) map.removeLayer(bases[k]); });
                map.addLayer(padrao);
                document.querySelectorAll('input[name="gp-base"]').forEach(function(r){
                    r.addEventListener('change', function(){
                        Object.keys(bases).forEach(function(k){ if (map.hasLayer(bases[k])) map.removeLayer(bases[k]); });
                        map.addLayer(bases[this.value]);
                    });
                });

                // --- Camadas (folhas: formato, situações dos trechos, SRE, contexto) ---
                var camadas = { {% for reg in this.regioes %}{% for it in reg.itens %}{% if it.situacoes %}{% for s in it.situacoes %}"{{ s.layer.get_name() }}": {{ s.layer.get_name() }}, {% endfor %}{% elif it.subitens %}{% for sc in it.subitens %}"{{ sc.layer.get_name() }}": {{ sc.layer.get_name() }}, {% endfor %}{% else %}"{{ it.layer.get_name() }}": {{ it.layer.get_name() }}, {% endif %}{% endfor %}{% endfor %}{% for c in this.contexto %}"{{ c.layer.get_name() }}": {{ c.layer.get_name() }}, {% endfor %} };

                // Polígonos de região: pano de fundo. Ao religar qualquer
                // camada o Leaflet a joga para o topo, então reempurramos as
                // regiões para trás depois de cada mudança.
                var camadasFundo = [ {% for reg in this.regioes %}{% for it in reg.itens %}{% if not it.situacoes and it.forma == 'poli' %}"{{ it.layer.get_name() }}", {% endif %}{% endfor %}{% endfor %} ];
                function regioesAoFundo() {
                    camadasFundo.forEach(function(k){
                        var l = camadas[k];
                        if (l && l.bringToBack && map.hasLayer(l)) l.bringToBack();
                    });
                }

                function aplicar(input) {
                    var lyr = camadas[input.getAttribute('data-camada')];
                    if (!lyr) return;
                    if (input.checked) {
                        if (!map.hasLayer(lyr)) map.addLayer(lyr);
                        regioesAoFundo();
                    } else {
                        if (map.hasLayer(lyr)) map.removeLayer(lyr);
                    }
                }

                // Apaga o bloco da região quando ela está toda desligada
                function pintarRegiao(rid) {
                    var mr = document.querySelector('#gp-painel input[data-regiao="'+rid+'"]:not([data-camada]):not([data-grupo])');
                    var bloco = document.querySelector('#gp-painel .gp-reg[data-reg="'+rid+'"]');
                    if (mr && bloco) bloco.classList.toggle('desligada', !mr.checked);
                }

                function sincronizarPais(rid, grp) {
                    if (grp) {
                        var f = document.querySelectorAll('#gp-painel input[data-camada][data-grupo="'+grp+'"]');
                        var g = document.querySelector('#gp-painel input[data-grupo="'+grp+'"]:not([data-camada])');
                        if (g) g.checked = Array.prototype.some.call(f, function(c){ return c.checked; });
                    }
                    if (rid) {
                        var fr = document.querySelectorAll('#gp-painel input[data-camada][data-regiao="'+rid+'"]');
                        var mr = document.querySelector('#gp-painel input[data-regiao="'+rid+'"]:not([data-camada]):not([data-grupo])');
                        if (mr) mr.checked = Array.prototype.some.call(fr, function(c){ return c.checked; });
                        pintarRegiao(rid);
                    }
                }

                // Botão "Desmarcar todas" / "Marcar todas"
                // (nome distinto de btnLimpar, que é o "x" da busca)
                var btnLimparReg = document.getElementById('gp-limpar-regioes');
                function atualizarBotaoLimpar() {
                    var todos = document.querySelectorAll('#gp-painel input[data-camada][data-regiao]');
                    var algum = Array.prototype.some.call(todos, function(c){ return c.checked; });
                    btnLimparReg.textContent = algum ? 'Desmarcar todas' : 'Marcar todas';
                    return algum;
                }
                btnLimparReg.addEventListener('click', function(){
                    var ligar = !atualizarBotaoLimpar();   // se nada ligado -> liga tudo
                    document.querySelectorAll('#gp-painel input[data-camada][data-regiao]').forEach(function(c){
                        c.checked = ligar; aplicar(c);
                    });
                    document.querySelectorAll('#gp-painel input[data-regiao]:not([data-camada])').forEach(function(m){
                        m.checked = ligar;
                    });
                    document.querySelectorAll('#gp-painel .gp-reg').forEach(function(b){
                        b.classList.toggle('desligada', !ligar);
                    });
                    atualizarBotaoLimpar();
                    setTimeout(sincronizarChips, 0);
                });

                // Estado inicial: sincroniza o mapa com os interruptores
                document.querySelectorAll('#gp-painel input[data-camada]').forEach(aplicar);

                // Nível 3 / folhas
                document.querySelectorAll('#gp-painel input[data-camada]').forEach(function(t){
                    t.addEventListener('change', function(){
                        aplicar(this);
                        sincronizarPais(this.getAttribute('data-regiao'), this.getAttribute('data-grupo'));
                    });
                });

                // Nível 2: grupo "Trechos" -> liga/desliga todas as situações
                document.querySelectorAll('#gp-painel input[data-grupo]:not([data-camada])').forEach(function(g){
                    g.addEventListener('change', function(){
                        var grp = this.getAttribute('data-grupo'), on = this.checked;
                        document.querySelectorAll('#gp-painel input[data-camada][data-grupo="'+grp+'"]').forEach(function(c){
                            c.checked = on; aplicar(c);
                        });
                        sincronizarPais(this.getAttribute('data-regiao'), null);
                    });
                });

                // Voa até um trecho (GeoJson) ou ponto crítico (marcador) e
                // abre a ficha dele.
                function irPara(bstr, fgNome, sre) {
                    var p = bstr.split(',').map(Number);
                    var ponto = (p[0] === p[2] && p[1] === p[3]);   // bounds degenerado = ponto
                    map.fitBounds([[p[0], p[1]], [p[2], p[3]]],
                                  { maxZoom: ponto ? 16 : 15, padding: [50, 50] });
                    var fg = camadas[fgNome];
                    if (!fg || !map.hasLayer(fg)) return;
                    fg.eachLayer(function(l){
                        // Ponto crítico: marcador solto -> casa pela coordenada
                        if (l.getLatLng && !l.eachLayer) {
                            var ll = l.getLatLng();
                            if (Math.abs(ll.lat - p[0]) < 1e-5 && Math.abs(ll.lng - p[1]) < 1e-5 && l.openPopup) {
                                setTimeout(function(){ l.openPopup(); }, 400);
                            }
                            return;
                        }
                        // Trecho: GeoJson -> casa pelo código SRE
                        if (!l.eachLayer) return;
                        l.eachLayer(function(sub){
                            if (sub.feature && sub.feature.properties &&
                                String(sub.feature.properties.SRE) === sre && sub.openPopup) {
                                setTimeout(function(){ sub.openPopup(); }, 350);
                            }
                        });
                    });
                }

                // Nível 4: clique no SRE da árvore
                document.querySelectorAll('#gp-painel .gp-sre').forEach(function(el){
                    el.addEventListener('click', function(e){
                        e.stopPropagation();
                        irPara(this.getAttribute('data-b'), this.getAttribute('data-fg'),
                               this.getAttribute('data-sre'));
                    });
                });

                // ------------------------------------------------ BUSCA
                var indice = [
                {%- for reg in this.regioes %}{%- for it in reg.itens %}{%- if it.situacoes %}{%- for s in it.situacoes %}{%- for sr in s.sres %}
                {s:"{{ sr.sre }}",r:"{{ sr.rod }}",c:"{{ sr.cid }}",g:"{{ reg.nome }}",t:"{{ s.codigo }}",k:"{{ s.cor }}",b:"{{ sr.b }}",f:"{{ s.layer.get_name() }}"},
                {%- endfor %}{%- endfor %}{%- elif it.subitens %}{%- for sc in it.subitens %}{%- for sr in sc.busca %}
                {s:"{{ sr.sre }}",r:"{{ sr.rod }}",c:"{{ sr.cid }}",g:"{{ reg.nome }}",t:"OAE",k:"{{ sc.cor }}",b:"{{ sr.b }}",f:"{{ sc.layer.get_name() }}"},
                {%- endfor %}{%- endfor %}{%- endif %}{%- endfor %}{%- endfor %}
                ];
                var campoBusca = document.getElementById('gp-busca');
                var caixaRes = document.getElementById('gp-resultados');
                var btnLimpar = document.getElementById('gp-limpar');

                function normalizar(t) {
                    return (t || '').toString().toLowerCase()
                        .normalize('NFD').replace(/[̀-ͯ]/g, '');
                }

                // Enquadra TODOS os resultados da busca (fly-to)
                var ultimosAchados = [];
                function enquadrarResultados() {
                    if (!ultimosAchados.length) return;
                    var lats = [], lngs = [];
                    ultimosAchados.forEach(function(i){
                        var p = i.b.split(',').map(Number);
                        lats.push(p[0], p[2]); lngs.push(p[1], p[3]);
                    });
                    map.flyToBounds([[Math.min.apply(null, lats), Math.min.apply(null, lngs)],
                                     [Math.max.apply(null, lats), Math.max.apply(null, lngs)]],
                                    { padding: [55, 55], maxZoom: 14, duration: 0.8 });
                }

                function renderResultados(termo) {
                    caixaRes.innerHTML = '';
                    btnLimpar.style.display = termo ? 'block' : 'none';
                    ultimosAchados = [];
                    if (!termo) return;
                    var q = normalizar(termo);
                    var achados = indice.filter(function(i){
                        return normalizar(i.s).indexOf(q) >= 0 ||
                               normalizar(i.r).indexOf(q) >= 0 ||
                               normalizar(i.c).indexOf(q) >= 0;
                    });
                    if (!achados.length) {
                        caixaRes.innerHTML = '<div class="gp-vazio">Nada encontrado</div>';
                        return;
                    }
                    ultimosAchados = achados;
                    var total = achados.length;

                    // Cabeçalho: quantos achou + enquadrar todos no mapa
                    var cab = document.createElement('div');
                    cab.className = 'gp-res-cab';
                    cab.innerHTML = '<span>' + total + (total > 1 ? ' resultados' : ' resultado') + '</span>';
                    var bt = document.createElement('button');
                    bt.type = 'button'; bt.className = 'gp-enquadrar';
                    bt.textContent = total > 1 ? 'ver todos no mapa' : 'ver no mapa';
                    bt.addEventListener('click', enquadrarResultados);
                    cab.appendChild(bt);
                    caixaRes.appendChild(cab);
                    achados.slice(0, 40).forEach(function(i){
                        var d = document.createElement('div');
                        d.className = 'gp-res';
                        d.innerHTML = '<span class="gp-res-txt">' +
                            '<span class="gp-res-sre">' + i.s + '</span>' +
                            '<span class="gp-res-sub">' + [i.r, i.c, i.g].filter(Boolean).join(' · ') + '</span>' +
                            '</span><span class="gp-res-tag" style="--c:' + i.k + '">' + i.t + '</span>';
                        d.addEventListener('click', function(){ irPara(i.b, i.f, i.s); });
                        caixaRes.appendChild(d);
                    });
                    if (total > 40) {
                        var m2 = document.createElement('div');
                        m2.className = 'gp-vazio';
                        m2.textContent = '… e mais ' + (total - 40) + ' resultado(s)';
                        caixaRes.appendChild(m2);
                    }
                }

                campoBusca.addEventListener('input', function(){ renderResultados(this.value.trim()); });
                // Enter = enquadra tudo que foi encontrado (não voa a cada tecla,
                // o que seria enjoativo enquanto se digita)
                campoBusca.addEventListener('keydown', function(e){
                    if (e.key === 'Enter') { e.preventDefault(); enquadrarResultados(); }
                });
                btnLimpar.addEventListener('click', function(){
                    campoBusca.value = ''; renderResultados(''); campoBusca.focus();
                });

                // ------------------------------------------------ FILTROS (chips)
                // Um chip pode mirar a situação (data-sit) ou o tipo de
                // inventário (data-inv); em ambos os casos vale para TODAS as
                // regiões de uma vez.
                function alvoDoChip(chip) {
                    return chip.hasAttribute('data-inv')
                        ? '#gp-painel input[data-inv="' + chip.getAttribute('data-inv') + '"]'
                        : '#gp-painel input[data-sit="' + chip.getAttribute('data-sit') + '"]';
                }
                document.querySelectorAll('#gp-painel .gp-chip').forEach(function(chip){
                    chip.addEventListener('click', function(){
                        var ligar = !this.classList.contains('on');
                        this.classList.toggle('on', ligar);
                        document.querySelectorAll(alvoDoChip(this)).forEach(function(c){
                            c.checked = ligar;
                            aplicar(c);
                            sincronizarPais(c.getAttribute('data-regiao'), c.getAttribute('data-grupo'));
                        });
                        setTimeout(pintarTudo, 0);
                    });
                });

                // Mantém o chip coerente quando a situação é mexida pela árvore
                function sincronizarChips() {
                    document.querySelectorAll('#gp-painel .gp-chip').forEach(function(chip){
                        var alvos = document.querySelectorAll(alvoDoChip(chip));
                        var algum = Array.prototype.some.call(alvos, function(c){ return c.checked; });
                        chip.classList.toggle('on', algum);
                    });
                }
                document.querySelectorAll('#gp-painel input[type="checkbox"]').forEach(function(c){
                    c.addEventListener('change', function(){
                        setTimeout(function(){
                            sincronizarChips(); atualizarBotaoLimpar(); pintarTudo();
                        }, 0);
                    });
                });
                document.querySelectorAll('#gp-painel .gp-reg').forEach(function(b){
                    pintarRegiao(b.getAttribute('data-reg'));
                });
                atualizarBotaoLimpar();
                pintarTudo();
                regioesAoFundo();

                // Nível 1: mestre da região -> liga/desliga tudo da região
                document.querySelectorAll('#gp-painel input[data-regiao]:not([data-camada]):not([data-grupo])').forEach(function(mst){
                    mst.addEventListener('change', function(){
                        var rid = this.getAttribute('data-regiao'), on = this.checked;
                        document.querySelectorAll('#gp-painel input[data-camada][data-regiao="'+rid+'"]').forEach(function(c){
                            c.checked = on; aplicar(c);
                        });
                        document.querySelectorAll('#gp-painel input[data-grupo][data-regiao="'+rid+'"]:not([data-camada])').forEach(function(g){
                            g.checked = on;
                        });
                        pintarRegiao(rid);   // apaga (ou reacende) o bloco da região
                    });
                });
            })();
        {% endmacro %}
        """)


def create_webgis():
    print("Iniciando a criação do WebGIS...")

    base_dir = Path(__file__).parent
    camadas_dir = base_dir / "camadas"
    output_file = base_dir / "mapa_interativo.html"

    # Falhas devolvem código != 0 para o atualizar.bat abortar antes de
    # publicar. Sem isso, um erro aqui passaria batido e o site iria ao ar
    # com o mapa antigo, sem ninguém perceber.
    if not camadas_dir.exists():
        print(f"ERRO: a pasta '{camadas_dir}' não existe. Mapa NÃO gerado.")
        sys.exit(1)

    m = folium.Map(location=[-15.7801, -47.9292], zoom_start=4, tiles=None,
                   max_zoom=21, zoom_control=False)

    # 1. Mapas de Fundo (o último adicionado abre por padrão -> Satélite)
    tl_padrao = folium.TileLayer('OpenStreetMap', name='Padrão', control=False,
                                 max_zoom=21, max_native_zoom=19)
    tl_escuro = folium.TileLayer(
        tiles='https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png',
        attr='&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a> contributors &copy; <a href="https://carto.com/attributions">CARTO</a>',
        name='Escuro', control=False, max_zoom=21, max_native_zoom=19)
    tl_satelite = folium.TileLayer(
        tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
        attr='Tiles &copy; Esri &mdash; Source: Esri, i-cubed, USDA, USGS, AEX, GeoEye, Getmapping, Aerogrid, IGN, IGP, UPR-EGP, and the GIS User Community',
        name='Satélite', control=False, max_zoom=21, max_native_zoom=17)
    for tl in (tl_padrao, tl_escuro, tl_satelite):
        tl.add_to(m)

    # 2. Ler shapefiles
    controle = carregar_controle_pontos()
    regioes = {}
    contexto = {}
    bounds = []
    total_situacao = {}

    for shp_file in sorted(camadas_dir.glob('*.shp')):
        nome = shp_file.stem
        tipo = classificar(nome)

        if tipo == 'skip':
            print(f"Ignorando (duplicata/base bruta): {shp_file.name}")
            continue
        if tipo is None:
            print(f"Aviso: tipo não reconhecido para '{nome}'. Ignorada.")
            continue

        mm = re.match(r'^(R\d+)_', nome.upper())
        rid = mm.group(1) if mm else 'OUTROS'
        if tipo in TIPO_ORDEM and rid in regioes and tipo in regioes[rid]:
            print(f"Ignorando duplicata: {shp_file.name} ({rid}/{tipo})")
            continue

        print(f"Lendo camada: {shp_file.name}")
        try:
            gdf = gpd.read_file(shp_file)

            if gdf.crs and gdf.crs.to_string() != 'EPSG:4326':
                gdf = gdf.to_crs(epsg=4326)
            elif not gdf.crs:
                print(f"  Aviso: {shp_file.name} sem CRS. Assumindo EPSG:4326.")
                gdf.set_crs(epsg=4326, inplace=True)

            gdf = limpar_atributos(gdf)

            # ---------------------------------------------- camadas de contexto
            if tipo == 'hidrografia':
                antes = len(gdf)
                if 'nuordemcda' in gdf.columns:
                    gdf = gdf[gdf['nuordemcda'] <= HIDRO_ORDEM_MAX].copy()
                gdf['geometry'] = gdf.geometry.simplify(HIDRO_SIMPLIFY, preserve_topology=False)
                print(f"  Hidrografia filtrada (ordem <= {HIDRO_ORDEM_MAX}): "
                      f"{antes} -> {len(gdf)} feições (simplificada)")
                fg = folium.FeatureGroup(name='Hidrografia', show=False, control=False)
                folium.GeoJson(data=gdf[['geometry']],
                               style_function=lambda x: {'color': COR_HIDRO, 'weight': 1.2,
                                                         'opacity': 0.75}).add_to(fg)
                contexto['hidrografia'] = {'fg': fg, 'count': len(gdf)}
                continue

            if tipo == 'estado':
                fg = folium.FeatureGroup(name='Limite do Estado', show=True, control=False)
                folium.GeoJson(data=gdf[['geometry']],
                               style_function=lambda x: {'color': COR_ESTADO, 'weight': 2.5,
                                                         'opacity': 0.9, 'fill': False,
                                                         'dashArray': '8,5'}).add_to(fg)
                contexto['estado'] = {'fg': fg, 'count': len(gdf)}
                continue

            # ------------------------------------------- inventário
            if tipo == 'inventario':
                mm = re.match(r'^(.+)_R(\d+)$', nome, re.IGNORECASE)
                chave = mm.group(1).upper()
                rid = 'R' + mm.group(2)
                if chave in INVENTARIO_APELIDOS:
                    certo = INVENTARIO_APELIDOS[chave]
                    print(f"  [NOME] '{nome}' parece erro de digitação de "
                          f"'{certo}_{rid}'. Tratando como {certo} — vale renomear o arquivo.")
                    chave = certo
                meta = next((m for m in INVENTARIO_META if m[0] == chave), None)
                if meta is None:
                    print(f"  Aviso: tipo de inventário '{chave}' desconhecido. Ignorado.")
                    continue
                _, rotulo, cor_inv, forma = meta

                # Descarta geometrias com coordenada inválida (NaN/fora de faixa)
                n_antes = len(gdf)
                gdf = gdf[gdf.geometry.apply(geom_valida)].copy()
                n_ruins = n_antes - len(gdf)
                if n_ruins:
                    print(f"  [QUALIDADE] {nome}: {n_ruins} feição(ões) com coordenada "
                          f"inválida descartada(s).")
                if len(gdf) == 0:
                    print(f"  Aviso: {nome} ficou sem feições válidas. Ignorada.")
                    continue

                minx, miny, maxx, maxy = gdf.total_bounds
                bounds.append([[miny, minx], [maxy, maxx]])

                cols = colunas_uteis(gdf)
                # Camada começa DESLIGADA (show=False)
                fg = folium.FeatureGroup(name=f'{rid}_inv_{chave}', show=False, control=False)

                # Só as pontes entram na busca: são poucas (centenas) e cada
                # uma é uma estrutura nomeada. Bueiros e descidas somam
                # milhares e pesariam o índice sem ganho real.
                busca_inv = []
                col_sre_inv = next((c for c in gdf.columns if str(c).upper() == 'SRE'), None)
                col_km_inv = next((c for c in gdf.columns if str(c).upper() == 'KM'), None)

                if forma == 'ponto':
                    # Pontos densos: agrupados em cluster (nao travam o mapa)
                    cluster = MarkerCluster(options={'maxClusterRadius': 45,
                                                     'disableClusteringAtZoom': 17})
                    for _, linha in gdf.iterrows():
                        if linha.geometry is None or linha.geometry.is_empty:
                            continue
                        campos = ''.join(
                            f'<tr><td style="color:#94a3b8;padding:1px 8px 1px 0">{_html.escape(str(c))}</td>'
                            f'<td style="font-weight:600">{_html.escape(str(linha[c]))}</td></tr>'
                            for c in cols if linha[c] is not None and str(linha[c]).strip() not in ('', 'nan', 'None'))
                        html = (f'<div style="font-family:Inter,sans-serif;font-size:11px">'
                                f'<div style="font-weight:700;color:{cor_inv};margin-bottom:4px">{_html.escape(rotulo)}</div>'
                                f'<table>{campos}</table></div>')
                        folium.CircleMarker(
                            location=[linha.geometry.y, linha.geometry.x],
                            radius=4, color='#ffffff', weight=1,
                            fill_color=cor_inv, fill_opacity=0.95,
                            popup=folium.Popup(html, max_width=280),
                        ).add_to(cluster)

                        if chave == 'OAE':
                            sre_v = js_safe(linha[col_sre_inv]) if col_sre_inv else ''
                            km_v = js_safe(linha[col_km_inv]) if col_km_inv else ''
                            busca_inv.append({
                                'sre': f'Ponte {sre_v}'.strip() if sre_v else 'Ponte',
                                'rod': f'km {km_v}' if km_v else '',
                                'cid': sre_v,     # permite achar pelo código SRE
                                'b': (f'{linha.geometry.y:.6f},{linha.geometry.x:.6f},'
                                      f'{linha.geometry.y:.6f},{linha.geometry.x:.6f}'),
                            })
                    cluster.add_to(fg)
                else:
                    # Linhas: GeoJson direto (linhas nao agrupam)
                    popup = (folium.GeoJsonPopup(fields=cols, aliases=cols, localize=True)
                             if cols else None)
                    folium.GeoJson(
                        data=gdf[cols + ['geometry']], popup=popup,
                        style_function=(lambda x, c=cor_inv: {'color': c, 'weight': 3, 'opacity': 0.85}),
                    ).add_to(fg)

                regioes.setdefault(rid, {}).setdefault('inventario', {'itens': {}})
                regioes[rid]['inventario']['itens'][chave] = {
                    'fg': fg, 'count': len(gdf), 'rotulo': rotulo,
                    'cor': cor_inv, 'forma': forma, 'busca': busca_inv,
                }
                print(f"  '{nome}' -> {rid} / inventário / {chave} ({len(gdf)} feições)")
                continue

            # ------------------------------------------- dados por região
            minx, miny, maxx, maxy = gdf.total_bounds
            bounds.append([[miny, minx], [maxy, maxx]])

            validar_nomenclatura_sre(gdf, nome)

            cols = [c for c in gdf.columns if c.lower() != 'geometry']
            popup = folium.GeoJsonPopup(fields=cols, aliases=cols, localize=True) if cols else None
            _, cor, _ = TIPO_META[tipo]
            if tipo == 'regiao':
                cor = cor_regiao(rid)   # cada região com a sua cor

            if tipo == 'trechos':
                # Uma camada por SITUAÇÃO: permite ligar/desligar cada classe
                csit = coluna_situacao(gdf)
                gdf['SITUACAO'] = (gdf[csit].astype(str).str.strip().str.upper()
                                   if csit else '')
                gdf['SITUACAO'] = gdf['SITUACAO'].replace({'NONE': '—', '': '—'})
                col_sre = next((c for c in gdf.columns if str(c).upper() == 'SRE'), None)
                col_rod = next((c for c in gdf.columns if str(c).upper() == 'RODOVIA'), None)
                col_cid = next((c for c in gdf.columns if 'CIDADE' in str(c).upper()), None)
                situacoes = []
                for sit in ordenar_situacoes(gdf['SITUACAO'].unique()):
                    sub = gdf[gdf['SITUACAO'] == sit]
                    desc_sit, cor_sit = info_situacao(sit)
                    cols_sub = [c for c in sub.columns if c.lower() != 'geometry']
                    fg_sit = folium.FeatureGroup(name=f'{rid}_trechos_{sit}',
                                                 show=True, control=False)
                    folium.GeoJson(
                        data=sub, name=f'{nome}_{sit}',
                        popup=folium.GeoJsonPopup(fields=cols_sub, aliases=cols_sub, localize=True),
                        style_function=(lambda x, c=cor_sit: {'color': c, 'weight': 3, 'opacity': 0.9}),
                    ).add_to(fg_sit)

                    # 4º nível: lista de SREs com o retângulo (bounds) de cada
                    # trecho, para o clique no código voar até ele no mapa.
                    lista_sre = []
                    for _, row in sub.iterrows():
                        if row.geometry is None or row.geometry.is_empty:
                            continue
                        minx_t, miny_t, maxx_t, maxy_t = row.geometry.bounds
                        lista_sre.append({
                            'sre': js_safe(row[col_sre]) if col_sre else '—',
                            'rod': js_safe(row[col_rod]) if col_rod else '',
                            'cid': js_safe(row[col_cid]) if col_cid else '',
                            'b': f'{miny_t:.6f},{minx_t:.6f},{maxy_t:.6f},{maxx_t:.6f}',
                        })
                    lista_sre.sort(key=lambda s: s['sre'])

                    situacoes.append({'codigo': sit, 'desc': desc_sit, 'fg': fg_sit,
                                      'count': len(sub), 'cor': cor_sit, 'sres': lista_sre})
                    total_situacao[sit] = total_situacao.get(sit, 0) + len(sub)
                regioes.setdefault(rid, {})['trechos'] = {'count': len(gdf),
                                                          'situacoes': situacoes}
                print(f"  '{nome}' -> {rid} / trechos ({len(gdf)} feições em "
                      f"{len(situacoes)} situações: "
                      f"{', '.join(s['codigo'] + '=' + str(s['count']) for s in situacoes)})")
                continue

            # ---------------------------------------- pontos críticos
            # Estrutura espelhando os trechos:
            #   Pontos críticos > Status > número do ponto (clique = voa)
            if tipo == 'criticos':
                col_id = next((c for c in gdf.columns if str(c).lower() == 'id'), None)
                sem_ficha = 0
                por_status = {}   # rótulo -> {'cor', 'pontos': [...]}
                for _, linha in gdf.iterrows():
                    if linha.geometry is None or linha.geometry.is_empty:
                        continue
                    try:
                        ponto = int(float(linha[col_id])) if col_id else 0
                    except (TypeError, ValueError):
                        ponto = 0
                    info = controle.get((rid, ponto))
                    if info is None:
                        sem_ficha += 1
                        info = {'rodovia': None, 'trecho': None, 'final': None, 'historico': []}
                    rotulo, cor_pt = _status_pc(status_atual(info))
                    por_status.setdefault(rotulo, {'cor': cor_pt, 'pontos': []})['pontos'].append(
                        (ponto, linha.geometry.y, linha.geometry.x, info))

                grupos = []
                for rotulo in ORDEM_STATUS_PC:
                    if rotulo not in por_status:
                        continue
                    bloco = por_status[rotulo]
                    cor_pt = bloco['cor']
                    fg_st = folium.FeatureGroup(name=f'{rid}_criticos_{rotulo}',
                                                show=True, control=False)
                    lista = []
                    for ponto, lat, lon, info in sorted(bloco['pontos'], key=lambda t: t[0]):
                        folium.CircleMarker(
                            location=[lat, lon], radius=3.5, color='#ffffff', weight=1.2,
                            fill_color=cor_pt, fill_opacity=1,
                            # A classe leva o halo/brilho no CSS. O Leaflet recria
                            # o <path> ao religar a camada, então a animação de
                            # entrada toca de novo a cada ativação.
                            className='pc-mk ' + CLASSE_STATUS_PC.get(cor_pt, 'pc-c0'),
                            popup=folium.Popup(ficha_ponto_critico(rid, ponto, info), max_width=300),
                            tooltip=f'Ponto {ponto:03d}',
                        ).add_to(fg_st)
                        lista.append({'sre': f'Ponto {ponto:03d}',
                                      'rod': js_safe(info.get('rodovia')),
                                      'cid': '',
                                      'b': f'{lat:.6f},{lon:.6f},{lat:.6f},{lon:.6f}'})
                    grupos.append({'codigo': rotulo, 'desc': '', 'fg': fg_st,
                                   'count': len(lista), 'cor': cor_pt, 'sres': lista})

                regioes.setdefault(rid, {})['criticos'] = {'count': len(gdf), 'situacoes': grupos}
                if sem_ficha:
                    print(f"  [QUALIDADE] {rid}: {sem_ficha} ponto(s) crítico(s) sem ficha na planilha.")
                print(f"  '{nome}' -> {rid} / criticos ({len(gdf)} pontos em "
                      f"{len(grupos)} status: "
                      f"{', '.join(g['codigo'] + '=' + str(g['count']) for g in grupos)})")
                continue

            # Atenção: no folium o style_function sobrepõe as opções do
            # CircleMarker, então o estilo do nó tem que sair daqui.
            if tipo == 'sre':
                # Nó discreto: anel marinho fino com miolo branco.
                style_fn = (lambda x: {'color': COR_SRE, 'fillColor': '#ffffff',
                                       'weight': 1.1, 'fillOpacity': 1, 'opacity': 0.85})
                marcador = folium.CircleMarker(radius=2.5)
            elif tipo == 'regiao':
                # A região é pano de fundo: a classe 'gp-regiao' desliga o
                # pointer-events no CSS, então o polígono NÃO rouba os cliques
                # dos trechos e pontos que estão sobre ele. Sem popup, pelo
                # mesmo motivo (ele nunca abriria).
                style_fn = (lambda x, color=cor: {'color': color, 'fillColor': color,
                                                  'weight': 3, 'fillOpacity': 0.35,
                                                  'className': 'gp-regiao'})
                marcador = folium.CircleMarker(radius=6, color=cor, fill_color=cor,
                                               fill_opacity=0.9, weight=1)
                popup = None
            else:
                style_fn = (lambda x, color=cor: {'color': color, 'fillColor': color,
                                                  'weight': 3, 'fillOpacity': 0.35})
                marcador = folium.CircleMarker(radius=6, color=cor, fill_color=cor,
                                               fill_opacity=0.9, weight=1)

            fg = folium.FeatureGroup(name=f'{rid}_{tipo}', show=True, control=False)
            folium.GeoJson(
                data=gdf, name=nome, popup=popup, style_function=style_fn,
                marker=marcador
            ).add_to(fg)
            regioes.setdefault(rid, {})[tipo] = {'fg': fg, 'count': len(gdf), 'cor': cor}
            print(f"  '{nome}' -> {rid} / {tipo} ({len(gdf)} feições)")

        except Exception as e:
            print(f"  Erro ao processar {shp_file.name}: {e}")

    # Rede de segurança: pasta existe mas nada carregou (nomes fora do padrão,
    # arquivos corrompidos). Sem isto, geraríamos um mapa vazio e o
    # atualizar.bat o publicaria como se estivesse tudo bem.
    if not regioes:
        print("ERRO: nenhuma camada de região foi carregada. Mapa NÃO gerado.")
        print("      Verifique os nomes dos arquivos: R{numero}_{TIPO}.shp")
        print(f"      (ex.: R1_TRECHOS.shp). Pasta lida: {camadas_dir}")
        sys.exit(1)

    # Ordem de sobreposição: hidrografia -> estado -> polígonos -> linhas -> pontos
    if 'hidrografia' in contexto:
        contexto['hidrografia']['fg'].add_to(m)
    if 'estado' in contexto:
        contexto['estado']['fg'].add_to(m)

    def rid_key(rid):
        mm = re.match(r'R(\d+)$', rid)
        return (0, int(mm.group(1))) if mm else (1, rid)
    ordem_regioes = sorted(regioes, key=rid_key)

    for tipo in TIPO_ORDEM:
        for rid in ordem_regioes:
            item = regioes[rid].get(tipo)
            if not item:
                continue
            if item.get('situacoes'):        # trechos e pontos críticos
                for s in item['situacoes']:
                    s['fg'].add_to(m)
            else:
                item['fg'].add_to(m)

    # Inventário por cima (mas desligado por padrão)
    for rid in ordem_regioes:
        inv = regioes[rid].get('inventario')
        if inv:
            for chave, it in inv['itens'].items():
                it['fg'].add_to(m)

    if bounds:
        m.fit_bounds([[min(b[0][0] for b in bounds), min(b[0][1] for b in bounds)],
                      [max(b[1][0] for b in bounds), max(b[1][1] for b in bounds)]])

    # 3. Ferramentas de navegação
    Fullscreen(position='topright', title='Tela cheia',
               title_cancel='Sair da tela cheia', force_separate_button=True).add_to(m)
    MousePosition(position='bottomright', separator=' | ', prefix='Coordenadas:',
                  num_digits=6, lat_first=True).add_to(m)

    # 4. Painel
    def fmt(n):
        return '{:,}'.format(n).replace(',', '.')

    def nome_regiao(rid):
        mm = re.match(r'R(\d+)$', rid)
        return f'Região {int(mm.group(1))}' if mm else rid.title()

    regioes_info = []
    for rid in ordem_regioes:
        itens = []
        for tipo in TIPO_ORDEM:
            item = regioes[rid].get(tipo)
            if not item:
                continue
            label, cor, forma = TIPO_META[tipo]
            if item.get('situacoes'):        # trechos e pontos críticos
                itens.append({
                    'nome': label, 'cor': cor, 'forma': forma,
                    'count': fmt(item['count']), 'grupo': f'{rid}-{tipo}',
                    'situacoes': [{'codigo': s['codigo'], 'desc': s['desc'],
                                   'layer': s['fg'], 'cor': s['cor'],
                                   'count': fmt(s['count']), 'sres': s['sres']}
                                  for s in item['situacoes']],
                })
            else:
                itens.append({'nome': label, 'layer': item['fg'],
                              'cor': item.get('cor', cor),   # região traz a sua
                              'forma': forma, 'count': fmt(item['count']),
                              'situacoes': None})

        # Grupo "Inventário": expande em subcamadas (folhas simples), na ordem
        # oficial de INVENTARIO_META.
        inv = regioes[rid].get('inventario')
        if inv:
            subitens = []
            total_inv = 0
            for chave, _rot, _cor, _forma in INVENTARIO_META:
                it = inv['itens'].get(chave)
                if not it:
                    continue
                total_inv += it['count']
                subitens.append({'nome': it['rotulo'], 'layer': it['fg'],
                                 'chave': chave,           # para o filtro global
                                 'cor': it['cor'], 'forma': it['forma'],
                                 'count': fmt(it['count']),
                                 'busca': it.get('busca') or []})
            if subitens:
                itens.append({'nome': 'Inventário', 'cor': COR_MARCA, 'forma': 'inv',
                              'count': fmt(total_inv), 'grupo': f'{rid}-inventario',
                              'situacoes': None, 'subitens': subitens})

        if itens:
            regioes_info.append({'id': rid, 'nome': nome_regiao(rid),
                                 'cor': cor_regiao(rid), 'itens': itens})

    contexto_info = []
    if 'estado' in contexto:
        contexto_info.append({'nome': 'Limite do Estado', 'layer': contexto['estado']['fg'],
                              'cor': COR_ESTADO, 'forma': 'contorno', 'ativo': True,
                              'count': fmt(contexto['estado']['count'])})
    if 'hidrografia' in contexto:
        contexto_info.append({'nome': 'Hidrografia', 'layer': contexto['hidrografia']['fg'],
                              'cor': COR_HIDRO, 'forma': 'linha', 'ativo': False,
                              'count': fmt(contexto['hidrografia']['count'])})

    # 5. Arquivos para download (KML + Shapefile zipado)
    print("\nGerando os arquivos para download...")
    catalogo = exportar_dados(base_dir, camadas_dir)

    logo_uri = None
    logo_path = base_dir / 'logo' / 'LOGO RTA.png'
    if logo_path.exists():
        logo_uri = 'data:image/png;base64,' + base64.b64encode(logo_path.read_bytes()).decode('ascii')

    # ---------------- Filtros globais (valem para TODAS as regiões) ----------
    # Situação dos trechos
    filtros_situacao = []
    for cod in ordenar_situacoes(total_situacao.keys()):
        desc, cor = info_situacao(cod)
        filtros_situacao.append({'codigo': cod, 'desc': desc, 'cor': cor,
                                 'count': fmt(total_situacao[cod])})

    # Status dos pontos críticos (somando as regiões)
    tot_pc = {}
    for rid in ordem_regioes:
        item = regioes[rid].get('criticos')
        for s in (item or {}).get('situacoes', []):
            tot_pc[s['codigo']] = tot_pc.get(s['codigo'], 0) + s['count']
    filtros_criticos = [
        {'codigo': cod, 'desc': cod, 'cor': _status_pc(cod)[1], 'count': fmt(tot_pc[cod])}
        for cod in ORDEM_STATUS_PC if cod in tot_pc
    ]

    # Inventário por tipo (somando as regiões)
    tot_inv = {}
    for rid in ordem_regioes:
        for chave, it in (regioes[rid].get('inventario') or {}).get('itens', {}).items():
            tot_inv[chave] = tot_inv.get(chave, 0) + it['count']
    filtros_inventario = [
        {'codigo': chave, 'desc': rotulo, 'cor': cor, 'count': fmt(tot_inv[chave])}
        for chave, rotulo, cor, _f in INVENTARIO_META if chave in tot_inv
    ]

    m.add_child(PainelControle(
        basemaps=[{'nome': 'Padrão', 'layer': tl_padrao, 'icone': 'mapa'},
                  {'nome': 'Satélite', 'layer': tl_satelite, 'icone': 'satelite'},
                  {'nome': 'Escuro', 'layer': tl_escuro, 'icone': 'lua'}],
        default_base=tl_satelite,
        regioes=regioes_info,
        contexto=contexto_info,
        filtros_situacao=filtros_situacao,
        filtros_criticos=filtros_criticos,
        filtros_inventario=filtros_inventario,
        downloads=catalogo,
        subtitulo='WebGIS · Inventário Rodoviário — Tocantins',
        logo=logo_uri,
    ))

    m.save(output_file)
    print(f"\nMapa salvo em: {output_file}")
    print(f"Regiões: {ordem_regioes}")
    print(f"Situação (total): {dict(sorted(total_situacao.items(), key=lambda kv: -kv[1]))}")


if __name__ == "__main__":
    create_webgis()
